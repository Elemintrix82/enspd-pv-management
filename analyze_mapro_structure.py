"""
Analyse détaillée de la structure du fichier MAPRO_GIT5_SN_SEM1.xlsx
pour identifier les différences avec les fichiers fonctionnels
"""
import openpyxl
from openpyxl import load_workbook
import json

def analyze_mapro_structure():
    """Analyse complète de la structure du fichier MAPRO"""

    print("=" * 80)
    print("ANALYSE STRUCTURE MAPRO_GIT5_SN_SEM1.xlsx")
    print("=" * 80)
    print()

    # Charger le fichier MAPRO
    mapro_path = r"D:\RYDI_Group\ENSPD\MAPRO_GIT5_SN_SEM1.xlsx"
    wb_mapro = load_workbook(mapro_path)
    ws_mapro = wb_mapro.active

    # Charger un fichier de référence fonctionnel
    ref_path = r"D:\RYDI_Group\ENSPD\media\pv\PV_GRT4_SEM7_ALT.xlsx"
    wb_ref = load_workbook(ref_path)
    ws_ref = wb_ref.active

    print("FICHIER MAPRO:")
    print(f"  Path: {mapro_path}")
    print(f"  Feuille: {ws_mapro.title}")
    print(f"  Dimensions: {ws_mapro.max_row} lignes x {ws_mapro.max_column} colonnes")
    print()

    print("FICHIER RÉFÉRENCE (PV_GRT4):")
    print(f"  Path: {ref_path}")
    print(f"  Feuille: {ws_ref.title}")
    print(f"  Dimensions: {ws_ref.max_row} lignes x {ws_ref.max_column} colonnes")
    print()

    # Analyser les 20 premières lignes du fichier MAPRO
    print("=" * 80)
    print("ANALYSE DES EN-TÊTES MAPRO (20 premières lignes)")
    print("=" * 80)

    for row_idx in range(1, 21):
        print(f"\n[Ligne {row_idx}]")

        # Afficher les 15 premières colonnes
        for col_idx in range(1, 16):
            cell_value = ws_mapro.cell(row=row_idx, column=col_idx).value
            if cell_value:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                # Tronquer si trop long
                value_str = str(cell_value)[:50]
                print(f"  {col_letter}: {value_str}")

    print()
    print("=" * 80)
    print("ANALYSE DES EN-TÊTES RÉFÉRENCE (20 premières lignes)")
    print("=" * 80)

    for row_idx in range(1, 21):
        print(f"\n[Ligne {row_idx}]")

        # Afficher les 15 premières colonnes
        for col_idx in range(1, 16):
            cell_value = ws_ref.cell(row=row_idx, column=col_idx).value
            if cell_value:
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                value_str = str(cell_value)[:50]
                print(f"  {col_letter}: {value_str}")

    print()
    print("=" * 80)
    print("DÉTECTION DE LA LIGNE DE DÉBUT DES DONNÉES")
    print("=" * 80)

    # Chercher la première ligne avec un matricule (colonne B généralement)
    data_start_mapro = None
    for row_idx in range(1, 30):
        cell_value = ws_mapro.cell(row=row_idx, column=2).value
        if cell_value and isinstance(cell_value, str) and len(cell_value) >= 6:
            # Probable matricule
            if any(char.isdigit() for char in cell_value):
                data_start_mapro = row_idx
                print(f"MAPRO - Première ligne de données: {data_start_mapro}")
                print(f"  Matricule trouvé: {cell_value}")
                break

    data_start_ref = None
    for row_idx in range(1, 30):
        cell_value = ws_ref.cell(row=row_idx, column=2).value
        if cell_value and isinstance(cell_value, str) and len(cell_value) >= 6:
            if any(char.isdigit() for char in cell_value):
                data_start_ref = row_idx
                print(f"RÉFÉRENCE - Première ligne de données: {data_start_ref}")
                print(f"  Matricule trouvé: {cell_value}")
                break

    print()
    print("=" * 80)
    print("ANALYSE DES COLONNES FIXES")
    print("=" * 80)

    # Analyser les colonnes A, B, C, D, E
    if data_start_mapro:
        print(f"\nMAPRO - Ligne {data_start_mapro}:")
        for col_idx in range(1, 6):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            header_value = ws_mapro.cell(row=data_start_mapro - 1, column=col_idx).value
            data_value = ws_mapro.cell(row=data_start_mapro, column=col_idx).value
            print(f"  Col {col_letter}: Header='{header_value}', Data='{data_value}'")

    if data_start_ref:
        print(f"\nRÉFÉRENCE - Ligne {data_start_ref}:")
        for col_idx in range(1, 6):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            header_value = ws_ref.cell(row=data_start_ref - 1, column=col_idx).value
            data_value = ws_ref.cell(row=data_start_ref, column=col_idx).value
            print(f"  Col {col_letter}: Header='{header_value}', Data='{data_value}'")

    print()
    print("=" * 80)
    print("RECHERCHE DES CODES UE/ECUE DANS MAPRO")
    print("=" * 80)

    # Chercher les patterns de codes dans les en-têtes
    if data_start_mapro:
        header_row = data_start_mapro - 1
        print(f"\nAnalyse de la ligne {header_row} (en-tête):")

        ue_codes_found = []
        ecue_codes_found = []

        for col_idx in range(1, ws_mapro.max_column + 1):
            cell_value = ws_mapro.cell(row=header_row, column=col_idx).value
            if cell_value:
                value_str = str(cell_value).upper()
                col_letter = openpyxl.utils.get_column_letter(col_idx)

                # Chercher des patterns de codes
                if 'MAPRO' in value_str or 'EPD' in value_str:
                    print(f"  Col {col_letter}: {cell_value}")

                    if 'MAPRO' in value_str:
                        # Probable code UE ou ECUE
                        if len(value_str) <= 10:  # Code court = probable UE
                            ue_codes_found.append((col_letter, cell_value))
                        else:  # Code avec intitulé = probable ECUE
                            ecue_codes_found.append((col_letter, cell_value))

        print(f"\nCodes UE potentiels trouvés: {len(ue_codes_found)}")
        for col, code in ue_codes_found:
            print(f"  {col}: {code}")

        print(f"\nCodes ECUE potentiels trouvés: {len(ecue_codes_found)}")
        for col, code in ecue_codes_found:
            print(f"  {col}: {code}")

    print()
    print("=" * 80)
    print("RECHERCHE DES CODES UE/ECUE DANS RÉFÉRENCE")
    print("=" * 80)

    if data_start_ref:
        header_row_ref = data_start_ref - 1
        print(f"\nAnalyse de la ligne {header_row_ref} (en-tête):")

        ue_codes_ref = []
        ecue_codes_ref = []

        for col_idx in range(1, min(ws_ref.max_column + 1, 50)):  # Limiter à 50 colonnes
            cell_value = ws_ref.cell(row=header_row_ref, column=col_idx).value
            if cell_value:
                value_str = str(cell_value).upper()
                col_letter = openpyxl.utils.get_column_letter(col_idx)

                if 'EPDGIT' in value_str or 'EPDTCO' in value_str:
                    if len(value_str) <= 10:
                        ue_codes_ref.append((col_letter, cell_value))
                    else:
                        ecue_codes_ref.append((col_letter, cell_value))

        print(f"\nCodes UE trouvés: {len(ue_codes_ref)}")
        for col, code in ue_codes_ref[:10]:  # Afficher les 10 premiers
            print(f"  {col}: {code}")

        print(f"\nCodes ECUE trouvés: {len(ecue_codes_ref)}")
        for col, code in ecue_codes_ref[:10]:
            print(f"  {col}: {code}")

    print()
    print("=" * 80)
    print("ANALYSE MULTI-NIVEAUX DES EN-TÊTES")
    print("=" * 80)

    # Analyser plusieurs lignes d'en-têtes potentielles
    if data_start_mapro:
        print(f"\nMAPRO - En-têtes (lignes {data_start_mapro - 3} à {data_start_mapro - 1}):")

        for header_row in range(max(1, data_start_mapro - 3), data_start_mapro):
            print(f"\n  Ligne {header_row}:")
            for col_idx in range(4, min(20, ws_mapro.max_column + 1)):
                cell_value = ws_mapro.cell(row=header_row, column=col_idx).value
                if cell_value:
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    value_str = str(cell_value)[:40]
                    print(f"    {col_letter}: {value_str}")

    print()
    print("=" * 80)
    print("RÉSUMÉ DES DIFFÉRENCES IDENTIFIÉES")
    print("=" * 80)

    differences = []

    if data_start_mapro != data_start_ref:
        diff = f"Ligne de début données: MAPRO={data_start_mapro} vs REF={data_start_ref}"
        differences.append(diff)
        print(f"[DIFF] {diff}")

    if len(ue_codes_found) == 0:
        diff = "MAPRO: Aucun code UE détecté avec pattern 'EPDGIT'/'EPDTCO'"
        differences.append(diff)
        print(f"[CRITICAL] {diff}")

    print()
    print("=" * 80)
    print("ANALYSE TERMINÉE")
    print("=" * 80)

    return {
        'mapro': {
            'data_start_row': data_start_mapro,
            'ue_codes': ue_codes_found,
            'ecue_codes': ecue_codes_found,
            'max_row': ws_mapro.max_row,
            'max_column': ws_mapro.max_column,
        },
        'reference': {
            'data_start_row': data_start_ref,
            'ue_codes': ue_codes_ref,
            'ecue_codes': ecue_codes_ref,
            'max_row': ws_ref.max_row,
            'max_column': ws_ref.max_column,
        },
        'differences': differences
    }

if __name__ == "__main__":
    result = analyze_mapro_structure()

    # Sauvegarder le résultat
    with open('ANALYSE_MAPRO_STRUCTURE.json', 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print()
    print("Résultat sauvegardé dans: ANALYSE_MAPRO_STRUCTURE.json")
