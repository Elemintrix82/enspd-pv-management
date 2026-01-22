"""
Inventaire complet des feuilles du fichier MAPRO_GIT5_SN_SEM1.xlsx
"""
from openpyxl import load_workbook
import re

def inventorier_feuilles_mapro():
    """Analyse toutes les feuilles du fichier MAPRO"""

    mapro_path = r"D:\RYDI_Group\ENSPD\MAPRO_GIT5_SN_SEM1.xlsx"

    print("=" * 80)
    print("INVENTAIRE COMPLET - MAPRO_GIT5_SN_SEM1.xlsx")
    print("=" * 80)
    print()

    wb = load_workbook(mapro_path, data_only=True)

    print(f"Nombre total de feuilles: {len(wb.sheetnames)}")
    print(f"Noms des feuilles: {wb.sheetnames}")
    print()

    print("=" * 80)
    print("ANALYSE DÉTAILLÉE PAR FEUILLE")
    print("=" * 80)

    feuilles_pv = []
    feuilles_non_pv = []

    for idx, nom_feuille in enumerate(wb.sheetnames, 1):
        print(f"\n{'='*80}")
        print(f"FEUILLE {idx}: {nom_feuille}")
        print('='*80)

        ws = wb[nom_feuille]

        # Analyser la feuille
        info_feuille = {
            'numero': idx,
            'nom': nom_feuille,
            'dimensions': f"{ws.max_row} lignes x {ws.max_column} colonnes",
            'filiere': 'Non détecté',
            'niveau': 'Non détecté',
            'semestre': 'Non détecté',
            'annee': 'Non détecté',
            'regime': 'Non détecté',
            'nb_ue': 0,
            'nb_ecue': 0,
            'nb_etudiants': 0,
            'est_pv_valide': False
        }

        # Chercher les métadonnées dans les 20 premières lignes
        has_matricule = False
        has_ue_codes = False
        premier_matricule = None
        ligne_matricule = None

        for row in range(1, min(21, ws.max_row + 1)):
            for col in range(1, min(15, ws.max_column + 1)):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value:
                    cell_str = str(cell_value)
                    cell_upper = cell_str.upper()

                    # Détecter MATRICULE
                    if 'MATRICULE' in cell_upper:
                        has_matricule = True
                        ligne_matricule = row

                    # Détecter codes UE
                    if any(prefix in cell_upper for prefix in ['EPDGIT', 'EPDTCO', 'MPGIT', 'MPSSI', 'MAPRO']):
                        has_ue_codes = True

                    # Détecter filière
                    if 'FILIERE' in cell_upper or 'FILI' in cell_upper:
                        # Chercher dans les cellules suivantes
                        for c in range(col, min(col + 5, ws.max_column + 1)):
                            val = ws.cell(row=row, column=c).value
                            if val and len(str(val)) > 5:
                                info_feuille['filiere'] = str(val)[:100]
                                break

                    # Détecter niveau
                    if 'NIVEAU' in cell_upper:
                        for c in range(col, min(col + 5, ws.max_column + 1)):
                            val = ws.cell(row=row, column=c).value
                            if val:
                                match = re.search(r'([3-5]|L[3-5]|M[1-2])', str(val).upper())
                                if match:
                                    info_feuille['niveau'] = match.group(1)
                                    break

                    # Détecter semestre
                    if 'SEMESTRE' in cell_upper or cell_upper.strip().startswith('S'):
                        match = re.search(r'S\s*([1-9]|10)', cell_upper)
                        if match:
                            info_feuille['semestre'] = f"S{match.group(1)}"

                    # Détecter année académique
                    match_annee = re.search(r'20\d{2}[/-]20\d{2}', cell_str)
                    if match_annee:
                        info_feuille['annee'] = match_annee.group(0)

                    # Détecter régime (FI, ALT, SN)
                    if any(regime in cell_upper for regime in ['FORMATION INITIALE', 'FI', 'ALTERNANCE', 'ALT', 'SN']):
                        if 'ALTERNANCE' in cell_upper or 'ALT' in cell_upper:
                            info_feuille['regime'] = 'ALT'
                        elif 'SN' in cell_upper:
                            info_feuille['regime'] = 'SN'
                        elif 'FI' in cell_upper or 'FORMATION INITIALE' in cell_upper:
                            info_feuille['regime'] = 'FI'

        # Chercher le premier matricule (première ligne de données)
        if ligne_matricule:
            for row in range(ligne_matricule + 1, min(ligne_matricule + 50, ws.max_row + 1)):
                for col in range(1, 5):
                    val = ws.cell(row=row, column=col).value
                    if val:
                        val_str = str(val)
                        # Un matricule typique: 24G01883, 21G00038, etc.
                        if re.match(r'\d{2}G\d{5}', val_str):
                            premier_matricule = val_str
                            # Compter les étudiants
                            for r in range(row, ws.max_row + 1):
                                mat = ws.cell(row=r, column=col).value
                                if mat and re.match(r'\d{2}G\d{5}', str(mat)):
                                    info_feuille['nb_etudiants'] += 1
                                else:
                                    break
                            break
                if premier_matricule:
                    break

        # Compter les UE et ECUE (dans les lignes d'en-têtes)
        if ligne_matricule and ligne_matricule > 2:
            # Ligne UE typiquement 2-3 lignes avant MATRICULE
            ue_row = ligne_matricule - 3
            ecue_row = ligne_matricule - 2

            for col in range(1, ws.max_column + 1):
                ue_val = ws.cell(row=ue_row, column=col).value
                if ue_val:
                    ue_str = str(ue_val)
                    if any(prefix in ue_str for prefix in ['EPDGIT', 'EPDTCO', 'MPGIT', 'MPSSI']) and ':' in ue_str:
                        info_feuille['nb_ue'] += 1

                ecue_val = ws.cell(row=ecue_row, column=col).value
                if ecue_val:
                    ecue_str = str(ecue_val)
                    if '(' in ecue_str and ')' in ecue_str and any(prefix in ecue_str for prefix in ['EPDGIT', 'EPDTCO', 'MPGIT', 'MPSSI']):
                        info_feuille['nb_ecue'] += 1

        # Déterminer si c'est un PV valide
        info_feuille['est_pv_valide'] = (has_matricule and has_ue_codes and info_feuille['nb_etudiants'] > 0)

        # Afficher les informations
        print(f"Dimensions: {info_feuille['dimensions']}")
        print(f"Filière: {info_feuille['filiere']}")
        print(f"Niveau: {info_feuille['niveau']}")
        print(f"Semestre: {info_feuille['semestre']}")
        print(f"Année académique: {info_feuille['annee']}")
        print(f"Régime: {info_feuille['regime']}")
        print(f"Nombre d'UE: {info_feuille['nb_ue']}")
        print(f"Nombre d'ECUE: {info_feuille['nb_ecue']}")
        print(f"Nombre d'étudiants: {info_feuille['nb_etudiants']}")
        print(f"Premier matricule: {premier_matricule if premier_matricule else 'Non trouvé'}")
        print(f"Est un PV valide: {'OUI' if info_feuille['est_pv_valide'] else 'NON'}")

        if info_feuille['est_pv_valide']:
            feuilles_pv.append(info_feuille)
        else:
            feuilles_non_pv.append(info_feuille)

    # Résumé
    print("\n" + "=" * 80)
    print("RÉSUMÉ DE L'INVENTAIRE")
    print("=" * 80)
    print(f"\nNombre total de feuilles: {len(wb.sheetnames)}")
    print(f"Feuilles PV valides: {len(feuilles_pv)}")
    print(f"Feuilles non-PV: {len(feuilles_non_pv)}")

    if feuilles_pv:
        print("\n--- FEUILLES PV VALIDES ---")
        for f in feuilles_pv:
            print(f"{f['numero']}. {f['nom']}")
            print(f"   - {f['filiere']}")
            print(f"   - Niveau {f['niveau']}, {f['semestre']}, {f['regime']}")
            print(f"   - {f['nb_etudiants']} étudiants, {f['nb_ue']} UE, {f['nb_ecue']} ECUE")

    if feuilles_non_pv:
        print("\n--- FEUILLES NON-PV (IGNORÉES) ---")
        for f in feuilles_non_pv:
            print(f"{f['numero']}. {f['nom']}")
            print(f"   - Raison: Pas de structure PV détectée")

    print("\n" + "=" * 80)

    return {
        'total': len(wb.sheetnames),
        'pv_valides': feuilles_pv,
        'non_pv': feuilles_non_pv
    }

if __name__ == "__main__":
    inventorier_feuilles_mapro()
