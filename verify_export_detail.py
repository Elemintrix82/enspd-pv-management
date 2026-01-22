"""
Vérification détaillée du contenu de l'export Excel
"""
import openpyxl

print("="*80)
print("VERIFICATION DETAILLEE DE L'EXPORT EXCEL")
print("="*80)

# Charger le fichier exporté
wb = openpyxl.load_workbook("test_export_sans_filtres.xlsx")
ws = wb.active

# Vérifier les métadonnées (lignes 1-7)
print("\nMETADONNEES DU PV:")
print("-"*80)
for row in range(1, 7):
    col_a = ws.cell(row=row, column=1).value
    if col_a:
        print(f"  {col_a}")

# Vérifier les en-têtes (ligne 8-10)
print("\nEN-TETES:")
print("-"*80)

# Ligne 8 - UE
print("Ligne 8 (UE):")
ue_count = 0
for col in range(1, 80):
    val = ws.cell(row=8, column=col).value
    if val and ('EPDGIT' in str(val) or 'EPDTCO' in str(val)):
        ue_count += 1
        print(f"  Col {col}: {val}")

# Ligne 9 - ECUE
print("\nLigne 9 (ECUE) - Premiers ECUE:")
ecue_count = 0
for col in range(1, 80):
    val = ws.cell(row=9, column=col).value
    if val and ('EPDGIT' in str(val) or 'EPDTCO' in str(val) or 'SYNTHESE' in str(val).upper()):
        ecue_count += 1
        if ecue_count <= 5:  # Afficher les 5 premiers
            print(f"  Col {col}: {val}")

# Ligne 10 - Détails
print("\nLigne 10 (Détails colonnes) - 20 premières:")
for col in range(1, 21):
    val = ws.cell(row=10, column=col).value
    print(f"  Col {col}: {val}")

# Vérifier les données du premier étudiant (ligne 11)
print("\nPREMIER ETUDIANT (Ligne 11):")
print("-"*80)
print(f"N°: {ws.cell(row=11, column=1).value}")
print(f"Matricule: {ws.cell(row=11, column=2).value}")
print(f"Nom: {ws.cell(row=11, column=3).value}")

# Vérifier les notes du premier étudiant
print("\nNOTES DU PREMIER ETUDIANT:")
print("(Colonnes 6-12 = première ECUE)")
for col in range(6, 13):
    header = ws.cell(row=10, column=col).value
    value = ws.cell(row=11, column=col).value
    value_type = type(value).__name__ if value is not None else "None"
    print(f"  Col {col} [{header}]: {value} ({value_type})")

# Vérifier la synthèse générale du premier étudiant
print("\nSYNTHESE GENERALE DU PREMIER ETUDIANT:")
print("(Dernières colonnes)")
last_cols = [ws.max_column - 2, ws.max_column - 1, ws.max_column]
for col in last_cols:
    header = ws.cell(row=10, column=col).value
    value = ws.cell(row=11, column=col).value
    print(f"  Col {col} [{header}]: {value}")

# Chercher et afficher des exemples de cellules NULL
print("\nRECHERCHE DE CELLULES VIDES (NULL):")
print("-"*80)
null_examples = []
for row in range(11, 16):  # 5 premières lignes
    nom = ws.cell(row=row, column=3).value
    for col in range(4, 30):  # Colonnes de notes
        val = ws.cell(row=row, column=col).value
        if val is None:
            header = ws.cell(row=10, column=col).value
            null_examples.append(f"  Ligne {row} ({nom}) | Col {col} [{header}]: VIDE")
            if len(null_examples) >= 10:
                break
    if len(null_examples) >= 10:
        break

print(f"Exemples de cellules vides trouvees ({len(null_examples)}):")
for ex in null_examples[:10]:
    print(ex)

print("\n" + "="*80)
print("VERIFICATION TERMINEE")
print("="*80)

wb.close()
