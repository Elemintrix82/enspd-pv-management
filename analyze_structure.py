"""
Script pour analyser la structure complète UE/ECUE du fichier Excel
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

file_path = r"D:\RYDI_Group\ENSPD\PV_GRT4_SEM7_ALT.xlsx"

print("="*80)
print("ANALYSE STRUCTURE UE/ECUE")
print("="*80)

# Charger avec openpyxl
wb = openpyxl.load_workbook(file_path)
ws = wb.active

print("\n1. LIGNE 9 (UE) - Unités d'Enseignement")
print("-"*80)
for col_idx in range(1, 50):
    cell_val = ws.cell(row=9, column=col_idx).value
    if cell_val:
        col_letter = get_column_letter(col_idx)
        print(f"Colonne {col_letter} ({col_idx}): {cell_val}")

print("\n2. LIGNE 10 (ECUE) - Matières")
print("-"*80)
for col_idx in range(1, 50):
    cell_val = ws.cell(row=10, column=col_idx).value
    if cell_val:
        col_letter = get_column_letter(col_idx)
        print(f"Colonne {col_letter} ({col_idx}): {cell_val}")

print("\n3. LIGNE 11 (SOUS-COLONNES) - CC, EX, MOY, CA, DECISION")
print("-"*80)
for col_idx in range(1, 50):
    cell_val = ws.cell(row=11, column=col_idx).value
    if cell_val and str(cell_val).strip():
        col_letter = get_column_letter(col_idx)
        print(f"Colonne {col_letter} ({col_idx}): {cell_val}")

print("\n4. STRUCTURE PAR UE")
print("-"*80)

# Lire avec pandas
df = pd.read_excel(file_path, header=None, nrows=11)

# Analyser les UE (ligne 9, index 8)
ue_row = df.iloc[8]
ecue_row = df.iloc[9]
header_row = df.iloc[10]

current_ue = None
current_ecue = None
structure = []

for col_idx in range(5, 50):  # Commencer après les colonnes identité
    ue_val = ue_row.iloc[col_idx] if col_idx < len(ue_row) else None
    ecue_val = ecue_row.iloc[col_idx] if col_idx < len(ecue_row) else None
    header_val = header_row.iloc[col_idx] if col_idx < len(header_row) else None

    # Détecter nouvelle UE
    if pd.notna(ue_val) and 'EPDGIT' in str(ue_val):
        current_ue = str(ue_val).strip()
        print(f"\nUE: {current_ue}")

    # Détecter nouvelle ECUE
    if pd.notna(ecue_val) and 'EPDGIT' in str(ecue_val):
        current_ecue = str(ecue_val).strip()
        print(f"  - ECUE: {current_ecue}")
        structure.append({
            'ue': current_ue,
            'ecue': current_ecue,
            'col_start': col_idx
        })

    # Afficher les sous-colonnes
    if pd.notna(header_val) and str(header_val).strip():
        print(f"       Col {get_column_letter(col_idx+1)} ({col_idx}): {header_val}")

print("\n5. MAPPING COLONNES POUR EXTRACTION")
print("-"*80)

# Lire les données étudiants
df_students = pd.read_excel(file_path, header=10)
print(f"\nNombre total de colonnes dans le DataFrame: {len(df_students.columns)}")
print(f"Colonnes: {list(df_students.columns[:30])}")

# Identifier les patterns
cc_cols = [i for i, col in enumerate(df_students.columns) if 'CC' in str(col)]
ex_cols = [i for i, col in enumerate(df_students.columns) if 'EX' in str(col)]
moy_cols = [i for i, col in enumerate(df_students.columns) if 'MOY' in str(col)]
ca_cols = [i for i, col in enumerate(df_students.columns) if 'CA' in str(col)]
dec_cols = [i for i, col in enumerate(df_students.columns) if 'DECISION' in str(col)]

print(f"\nColonnes CC trouvées: {len(cc_cols)}")
print(f"Colonnes EX trouvées: {len(ex_cols)}")
print(f"Colonnes MOY trouvées: {len(moy_cols)}")
print(f"Colonnes CA trouvées: {len(ca_cols)}")
print(f"Colonnes DECISION trouvées: {len(dec_cols)}")

print("\n6. EXEMPLE: Première ligne étudiant")
print("-"*80)
first_student = df_students.iloc[0]

print(f"Étudiant: {first_student['NOMS & PRENOMS']}")
print(f"Matricule: {first_student['MATRICULE']}")

# Afficher toutes les notes
for i in range(len(cc_cols)):
    if i < len(cc_cols):
        cc_idx = cc_cols[i]
        ex_idx = ex_cols[i] if i < len(ex_cols) else None
        moy_idx = moy_cols[i] if i < len(moy_cols) else None
        ca_idx = ca_cols[i] if i < len(ca_cols) else None
        dec_idx = dec_cols[i] if i < len(dec_cols) else None

        print(f"\nNote #{i+1}:")
        print(f"  CC: {first_student.iloc[cc_idx] if cc_idx is not None else 'N/A'}")
        print(f"  EX: {first_student.iloc[ex_idx] if ex_idx is not None else 'N/A'}")
        print(f"  MOY: {first_student.iloc[moy_idx] if moy_idx is not None else 'N/A'}")
        print(f"  CA: {first_student.iloc[ca_idx] if ca_idx is not None else 'N/A'}")
        print(f"  DEC: {first_student.iloc[dec_idx] if dec_idx is not None else 'N/A'}")
