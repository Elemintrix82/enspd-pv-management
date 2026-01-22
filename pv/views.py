from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.contrib import messages
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.db.models import Q, Count
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from datetime import datetime
import tempfile
import os

from .models import ProcesVerbal, Etudiant, UE, ECUE, Note, SyntheseUE
from .forms import PVUploadForm
from .utils.excel_parser import PVExcelParser


def home(request):
    """
    Page d'accueil avec liste des PV importés
    """
    pvs = ProcesVerbal.objects.all()[:10]  # 10 derniers PV
    context = {
        'pvs': pvs,
        'total_pvs': ProcesVerbal.objects.count()
    }
    return render(request, 'pv/home.html', context)


def import_pv(request):
    """
    Vue pour importer un fichier PV Excel
    """
    if request.method == 'POST':
        form = PVUploadForm(request.POST, request.FILES)
        if form.is_valid():
            pv_instance = None
            try:
                # Sauvegarder temporairement le fichier
                uploaded_file = request.FILES['fichier']

                # Créer un fichier temporaire
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
                    for chunk in uploaded_file.chunks():
                        tmp_file.write(chunk)
                    tmp_file_path = tmp_file.name

                # Parser le fichier Excel
                parser = PVExcelParser(tmp_file_path)
                data = parser.parse()

                # Créer le ProcesVerbal avec transaction atomique
                with transaction.atomic():
                    # Sauvegarder le ProcesVerbal
                    pv_instance = form.save(commit=False)
                    pv_instance.filiere = data['metadata']['filiere']
                    pv_instance.niveau = data['metadata']['niveau']
                    pv_instance.semestre = data['metadata']['semestre']
                    pv_instance.annee_academique = data['metadata']['annee_academique']
                    pv_instance.formation = data['metadata'].get('formation', '')
                    pv_instance.save()

                    # Créer les UE
                    ue_objects = {}
                    for ue_data in data['ues']:
                        ue = UE.objects.create(
                            pv=pv_instance,
                            code=ue_data['code'],
                            intitule=ue_data['intitule'],
                            ordre=ue_data['ordre']
                        )
                        ue_objects[ue_data['code']] = ue

                    # Créer les ECUE (exclure les synthèses)
                    ecue_objects = {}
                    for ecue_data in data['ecues']:
                        # Ignorer les synthèses UE (elles ne sont pas des ECUE)
                        if ecue_data.get('is_synthese', False):
                            continue

                        ue_parent = ue_objects.get(ecue_data['ue_code'])
                        if ue_parent:
                            ecue = ECUE.objects.create(
                                ue=ue_parent,
                                code=ecue_data['code'],
                                intitule=ecue_data['intitule'],
                                ordre=ecue_data['ordre']
                            )
                            ecue_objects[ecue_data['code']] = ecue

                    # Créer les étudiants et leurs notes
                    for etudiant_data in data['etudiants']:
                        # Créer l'étudiant avec des valeurs nullables pour moyenne, crédits et décision
                        # Ces valeurs seront calculées après l'import des notes
                        etudiant = Etudiant.objects.create(
                            pv=pv_instance,
                            numero=etudiant_data['numero'],
                            matricule=etudiant_data['matricule'],
                            nom_prenom=etudiant_data['nom_prenom'],
                            moyenne_generale=etudiant_data.get('moyenne_generale'),
                            credits_acquis=etudiant_data.get('credits_acquis'),
                            decision_generale=etudiant_data.get('decision_generale')
                        )

                        # Créer les notes ECUE
                        for note_data in etudiant_data['notes']:
                            ecue = ecue_objects.get(note_data['ecue_code'])
                            if ecue:
                                Note.objects.create(
                                    etudiant=etudiant,
                                    ecue=ecue,
                                    cc=note_data.get('cc'),
                                    examen=note_data.get('examen'),
                                    moyenne=note_data.get('moyenne'),
                                    credit_attribue=note_data.get('credit_attribue'),
                                    decision=note_data.get('decision')
                                )

                        # Créer les synthèses UE
                        for synthese_data in etudiant_data.get('syntheses_ue', []):
                            ue = ue_objects.get(synthese_data['ue_code'])
                            if ue:
                                SyntheseUE.objects.create(
                                    etudiant=etudiant,
                                    ue=ue,
                                    moyenne_ue=synthese_data.get('moyenne_ue'),
                                    credits_attribues=synthese_data.get('credits_attribues'),
                                    decision=synthese_data.get('decision')
                                )

                        # Calculer et mettre à jour les résultats si les données Excel sont vides/nulles
                        if etudiant.moyenne_generale is None or etudiant.decision_generale is None:
                            etudiant.mettre_a_jour_resultats()

                    messages.success(
                        request,
                        f"✅ {pv_instance.nombre_etudiants} étudiants importés avec succès "
                        f"(Filière {pv_instance.filiere} - Niveau {pv_instance.niveau} - {pv_instance.semestre})"
                    )
                    return redirect('pv:dashboard', pk=pv_instance.pk)

            except Exception as e:
                messages.error(request, f"❌ Erreur lors de l'import: {str(e)}")
                # Si pv_instance a été sauvegardé (a un ID), on le supprime
                if pv_instance and pv_instance.pk:
                    try:
                        pv_instance.delete()
                    except Exception:
                        pass

    else:
        form = PVUploadForm()

    context = {'form': form}
    return render(request, 'pv/import.html', context)


def dashboard(request, pk):
    """
    Dashboard principal avec statistiques et tableau des étudiants
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Récupérer les paramètres de filtrage
    decision_filter = request.GET.get('decision', '')
    ue_filter = request.GET.get('ue', '')
    ecue_filter = request.GET.get('ecue', '')
    decision_ecue_filter = request.GET.get('decision_ecue', '')
    search_query = request.GET.get('search', '')
    moy_min = request.GET.get('moy_min', '')
    moy_max = request.GET.get('moy_max', '')

    # Filtrer les étudiants
    etudiants = pv.etudiants.all()

    if decision_filter:
        etudiants = etudiants.filter(decision_generale=decision_filter)

    if search_query:
        etudiants = etudiants.filter(
            Q(nom_prenom__icontains=search_query) |
            Q(matricule__icontains=search_query)
        )

    if moy_min:
        try:
            etudiants = etudiants.filter(moyenne_generale__gte=Decimal(moy_min))
        except:
            pass

    if moy_max:
        try:
            etudiants = etudiants.filter(moyenne_generale__lte=Decimal(moy_max))
        except:
            pass

    # Filtrage par ECUE spécifique
    if ecue_filter:
        if decision_ecue_filter:
            etudiants = etudiants.filter(
                notes__ecue__code=ecue_filter,
                notes__decision=decision_ecue_filter
            ).distinct()
        else:
            etudiants = etudiants.filter(notes__ecue__code=ecue_filter).distinct()

    # OPTIMISATION : Précharger les relations pour éviter les N+1 queries
    etudiants = etudiants.prefetch_related(
        'notes',
        'notes__ecue',
        'syntheses_ue',
        'syntheses_ue__ue'
    ).order_by('numero')

    # AMÉLIORATION 1 : Filtres dynamiques des colonnes
    # Récupérer les UE et ECUE pour l'affichage dynamique (filtrés selon les paramètres)
    if ecue_filter:
        # Si ECUE sélectionné : afficher UNIQUEMENT cet ECUE
        ecues = ECUE.objects.filter(code=ecue_filter, ue__pv=pv).prefetch_related('ue')
        ues = UE.objects.filter(ecues__in=ecues).distinct().prefetch_related('ecues').order_by('ordre')
    elif ue_filter:
        # Si UE sélectionnée : afficher UNIQUEMENT cette UE (avec tous ses ECUE)
        ues = pv.ues.filter(code=ue_filter).prefetch_related('ecues').order_by('ordre')
        ecues = ECUE.objects.filter(ue__in=ues).order_by('ordre')
    else:
        # Aucun filtre : afficher TOUTES les UE et ECUE
        ues = pv.ues.all().prefetch_related('ecues').order_by('ordre')
        ecues = ECUE.objects.filter(ue__pv=pv).order_by('ordre')

    # Pagination
    per_page = request.GET.get('per_page', 20)
    paginator = Paginator(etudiants, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    # Nouveau : Créer une structure pour afficher les notes par étudiant
    # On va organiser les données pour le template
    students_with_notes = []
    for etudiant in page_obj.object_list:
        # Récupérer toutes les notes de l'étudiant organisées par ECUE
        notes_par_ecue = {}
        for note in etudiant.notes.all():
            notes_par_ecue[note.ecue.code] = {
                'cc': note.cc,
                'examen': note.examen,
                'moyenne': note.moyenne,
                'credit_attribue': note.credit_attribue,
                'decision': note.decision,
                'decision_display': note.get_decision_display(),
                'decision_badge_class': note.get_decision_badge_class()
            }
        
        # Récupérer les synthèses UE
        syntheses_par_ue = {}
        for synthese in etudiant.syntheses_ue.all():
            syntheses_par_ue[synthese.ue.code] = {
                'moyenne_ue': synthese.moyenne_ue,
                'credits_attribues': synthese.credits_attribues,
                'decision': synthese.decision,
                'decision_display': synthese.get_decision_display(),
                'decision_badge_class': 'bg-info'  # Tu peux créer une méthode comme pour les notes
            }

        students_with_notes.append({
            'etudiant': etudiant,
            'notes_par_ecue': notes_par_ecue,
            'syntheses_par_ue': syntheses_par_ue
        })
        
    
    # CALCUL DES COLONNES POUR LE TABLEAU DÉTAILLÉ
    # Structure : 3 colonnes fixes + (pour chaque UE : ECUE*5 + 3 synthèse) + 3 synthèse générale
    
    total_colonnes_tableau = 3  # N°, Matricule, Nom & Prénom
    
    # Calculer le colspan pour chaque UE
    ues_with_colspan = []
    for ue in ues:
        # Nombre d'ECUE dans cette UE
        nb_ecues = ue.ecues.count()
        # Colonnes pour cette UE : (ECUE * 5) + 3 colonnes synthèse
        ue_colspan = (nb_ecues * 5) + 3
        ues_with_colspan.append({
            'ue': ue,
            'colspan': ue_colspan,
            'nb_ecues': nb_ecues
        })
        total_colonnes_tableau += ue_colspan
    
    total_colonnes_tableau += 3  # +3 pour synthèse générale

    context = {
        'pv': pv,
        'page_obj': page_obj,
        'etudiants': page_obj.object_list,
        'students_with_notes': students_with_notes,  # Nouveau
        'ues': ues,
        'ecues': ecues,
        'total_etudiants': etudiants.count(),
        'decision_filter': decision_filter,
        'ue_filter': ue_filter,
        'ecue_filter': ecue_filter,
        'decision_ecue_filter': decision_ecue_filter,
        'search_query': search_query,
        'moy_min': moy_min,
        'moy_max': moy_max,
        'per_page': per_page,
        'total_colonnes_tableau': total_colonnes_tableau,  # NOUVEAU
        'ues_with_colspan': ues_with_colspan,  # NOUVEAU : UE avec colspan calculé
    }

    return render(request, 'pv/dashboard.html', context)


def export_excel(request, pk):
    """
    Exporter les données filtrées en Excel avec notes détaillées
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Appliquer TOUS les mêmes filtres que le dashboard
    decision_filter = request.GET.get('decision', '')
    ue_filter = request.GET.get('ue', '')
    ecue_filter = request.GET.get('ecue', '')
    decision_ecue_filter = request.GET.get('decision_ecue', '')
    search_query = request.GET.get('search', '')
    moy_min = request.GET.get('moy_min', '')
    moy_max = request.GET.get('moy_max', '')

    etudiants = pv.etudiants.all()

    # Filtre par décision globale
    if decision_filter:
        etudiants = etudiants.filter(decision_generale=decision_filter)

    # Filtre par recherche
    if search_query:
        etudiants = etudiants.filter(
            Q(nom_prenom__icontains=search_query) |
            Q(matricule__icontains=search_query)
        )

    # Filtre par moyenne min/max
    if moy_min:
        try:
            etudiants = etudiants.filter(moyenne_generale__gte=Decimal(moy_min))
        except:
            pass

    if moy_max:
        try:
            etudiants = etudiants.filter(moyenne_generale__lte=Decimal(moy_max))
        except:
            pass

    # Filtre par ECUE spécifique
    if ecue_filter:
        if decision_ecue_filter:
            etudiants = etudiants.filter(
                notes__ecue__code=ecue_filter,
                notes__decision=decision_ecue_filter
            ).distinct()
        else:
            etudiants = etudiants.filter(notes__ecue__code=ecue_filter).distinct()

    # Précharger les relations
    etudiants = etudiants.prefetch_related(
        'notes',
        'notes__ecue',
        'syntheses_ue',
        'syntheses_ue__ue'
    ).order_by('numero')
    
    # Récupérer les UE et ECUE
    ues = pv.ues.all().prefetch_related('ecues').order_by('ordre')

    # Créer le fichier Excel avec notes détaillées
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PV Export"

    # Styles
    header_fill_blue = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_fill_info = PatternFill(start_color="17A2B8", end_color="17A2B8", fill_type="solid")
    header_fill_warning = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
    header_fill_success = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF")
    header_font_dark = Font(bold=True, color="000000")
    center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-tête avec informations du PV
    ws['A1'] = 'UNIVERSITÉ DE DOUALA'
    ws['A2'] = 'École Nationale Supérieure Polytechnique de Douala'
    ws['A3'] = f"Filière: {pv.filiere}"
    ws['A4'] = f"Niveau: {pv.niveau} - Semestre: {pv.semestre}"
    ws['A5'] = f"Année académique: {pv.annee_academique}"
    if pv.formation:
        ws['A6'] = f"Formation: {pv.formation}"

    # Ligne de départ pour le tableau
    header_start_row = 8
    current_col = 1

    # Ligne 1 : Colonnes fixes + En-têtes UE
    ws.merge_cells(start_row=header_start_row, start_column=1, end_row=header_start_row+2, end_column=1)
    cell = ws.cell(row=header_start_row, column=1, value="N°")
    cell.fill = header_fill_blue
    cell.font = header_font_white
    cell.alignment = center_alignment
    cell.border = border

    ws.merge_cells(start_row=header_start_row, start_column=2, end_row=header_start_row+2, end_column=2)
    cell = ws.cell(row=header_start_row, column=2, value="MATRICULE")
    cell.fill = header_fill_blue
    cell.font = header_font_white
    cell.alignment = center_alignment
    cell.border = border

    ws.merge_cells(start_row=header_start_row, start_column=3, end_row=header_start_row+2, end_column=3)
    cell = ws.cell(row=header_start_row, column=3, value="NOMS & PRÉNOMS")
    cell.fill = header_fill_blue
    cell.font = header_font_white
    cell.alignment = center_alignment
    cell.border = border

    current_col = 4

    # Pour chaque UE, créer les en-têtes
    for ue in ues:
        nb_ecues = ue.ecues.count()
        # Colonnes pour cette UE : (nb_ecues * 5) + 3 (synthèse UE)
        ue_colspan = (nb_ecues * 5) + 3
        ue_end_col = current_col + ue_colspan - 1

        # Ligne 1 : Nom de l'UE
        ws.merge_cells(start_row=header_start_row, start_column=current_col, end_row=header_start_row, end_column=ue_end_col)
        cell = ws.cell(row=header_start_row, column=current_col, value=f"{ue.code} - {ue.intitule}")
        cell.fill = header_fill_info
        cell.font = header_font_white
        cell.alignment = center_alignment
        cell.border = border

        # Ligne 2 : En-têtes ECUE
        ecue_col = current_col
        for ecue in ue.ecues.all():
            ws.merge_cells(start_row=header_start_row+1, start_column=ecue_col, end_row=header_start_row+1, end_column=ecue_col+4)
            cell = ws.cell(row=header_start_row+1, column=ecue_col, value=f"{ecue.code} ({ecue.credits} crédits)")
            cell.fill = header_fill_blue
            cell.font = header_font_white
            cell.alignment = center_alignment
            cell.border = border

            # Ligne 3 : CC, EX, MOY, CA, DEC
            for idx, label in enumerate(['CC', 'EX', 'MOY', 'CA', 'DEC']):
                cell = ws.cell(row=header_start_row+2, column=ecue_col+idx, value=label)
                cell.fill = header_fill_blue
                cell.font = header_font_white
                cell.alignment = center_alignment
                cell.border = border

            ecue_col += 5

        # Synthèse UE
        ws.merge_cells(start_row=header_start_row+1, start_column=ecue_col, end_row=header_start_row+1, end_column=ecue_col+2)
        cell = ws.cell(row=header_start_row+1, column=ecue_col, value=f"SYNTHÈSE UE {ue.code}")
        cell.fill = header_fill_warning
        cell.font = header_font_dark
        cell.alignment = center_alignment
        cell.border = border

        for idx, label in enumerate(['MOY', 'CRED', 'DEC']):
            cell = ws.cell(row=header_start_row+2, column=ecue_col+idx, value=label)
            cell.fill = header_fill_warning
            cell.font = header_font_dark
            cell.alignment = center_alignment
            cell.border = border

        current_col = ue_end_col + 1

    # Synthèse générale
    ws.merge_cells(start_row=header_start_row, start_column=current_col, end_row=header_start_row+1, end_column=current_col+2)
    cell = ws.cell(row=header_start_row, column=current_col, value="SYNTHÈSE GÉNÉRALE")
    cell.fill = header_fill_success
    cell.font = header_font_white
    cell.alignment = center_alignment
    cell.border = border

    for idx, label in enumerate(['MOY', 'CRED', 'DEC']):
        cell = ws.cell(row=header_start_row+2, column=current_col+idx, value=label)
        cell.fill = header_fill_success
        cell.font = header_font_white
        cell.alignment = center_alignment
        cell.border = border

    # Données des étudiants
    data_start_row = header_start_row + 3
    for row_idx, etudiant in enumerate(etudiants, start=data_start_row):
        # Colonnes fixes
        ws.cell(row=row_idx, column=1, value=etudiant.numero).alignment = center_alignment
        ws.cell(row=row_idx, column=2, value=etudiant.matricule).alignment = center_alignment
        ws.cell(row=row_idx, column=3, value=etudiant.nom_prenom)

        # Organiser les notes par ECUE code
        notes_dict = {note.ecue.code: note for note in etudiant.notes.all()}
        syntheses_dict = {synthese.ue.code: synthese for synthese in etudiant.syntheses_ue.all()}

        current_col = 4

        # Pour chaque UE
        for ue in ues:
            # Pour chaque ECUE
            for ecue in ue.ecues.all():
                note = notes_dict.get(ecue.code)
                if note:
                    # CC - Laisser vide si None
                    cell = ws.cell(row=row_idx, column=current_col)
                    cell.value = float(note.cc) if note.cc else None
                    cell.alignment = center_alignment

                    # EX - Laisser vide si None
                    cell = ws.cell(row=row_idx, column=current_col+1)
                    cell.value = float(note.examen) if note.examen else None
                    cell.alignment = center_alignment

                    # MOY - Laisser vide si None
                    cell = ws.cell(row=row_idx, column=current_col+2)
                    cell.value = float(note.moyenne) if note.moyenne else None
                    cell.alignment = center_alignment

                    # CA - Laisser vide si None ou 0
                    cell = ws.cell(row=row_idx, column=current_col+3)
                    cell.value = note.credit_attribue if note.credit_attribue else None
                    cell.alignment = center_alignment

                    # DEC - Laisser vide si None
                    cell = ws.cell(row=row_idx, column=current_col+4)
                    cell.value = note.decision if note.decision else None
                    cell.alignment = center_alignment
                else:
                    # Pas de note du tout - toutes les cellules vides
                    for i in range(5):
                        cell = ws.cell(row=row_idx, column=current_col+i)
                        cell.value = None
                        cell.alignment = center_alignment
                current_col += 5

            # Synthèse UE
            synthese = syntheses_dict.get(ue.code)
            if synthese:
                # MOY UE - Laisser vide si None
                cell = ws.cell(row=row_idx, column=current_col)
                cell.value = float(synthese.moyenne_ue) if synthese.moyenne_ue else None
                cell.alignment = center_alignment

                # CRED - Laisser vide si None ou 0
                cell = ws.cell(row=row_idx, column=current_col+1)
                cell.value = synthese.credits_attribues if synthese.credits_attribues else None
                cell.alignment = center_alignment

                # DEC - Laisser vide si None
                cell = ws.cell(row=row_idx, column=current_col+2)
                cell.value = synthese.decision if synthese.decision else None
                cell.alignment = center_alignment
            else:
                # Pas de synthèse - toutes les cellules vides
                for i in range(3):
                    cell = ws.cell(row=row_idx, column=current_col+i)
                    cell.value = None
                    cell.alignment = center_alignment
            current_col += 3

        # Synthèse générale
        ws.cell(row=row_idx, column=current_col, value=float(etudiant.moyenne_generale)).alignment = center_alignment
        ws.cell(row=row_idx, column=current_col+1, value=etudiant.credits_acquis).alignment = center_alignment
        ws.cell(row=row_idx, column=current_col+2, value=etudiant.get_decision_generale_display()).alignment = center_alignment

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 35
    for col_idx in range(4, current_col + 3):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 8
    
    # Générer la réponse HTTP
    filename = f"PV_{pv.filiere}_{pv.niveau}_{pv.semestre}_Export_{datetime.now().strftime('%Y%m%d')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def export_feuille_emargement(request, pk):
    """
    AMÉLIORATION 2 : Exporter la feuille d'émargement (liste simplifiée pour signatures)
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Appliquer TOUS les mêmes filtres que le dashboard
    decision_filter = request.GET.get('decision', '')
    ue_filter = request.GET.get('ue', '')
    ecue_filter = request.GET.get('ecue', '')
    decision_ecue_filter = request.GET.get('decision_ecue', '')
    search_query = request.GET.get('search', '')
    moy_min = request.GET.get('moy_min', '')
    moy_max = request.GET.get('moy_max', '')

    etudiants = pv.etudiants.all()

    # Filtre par décision globale
    if decision_filter:
        etudiants = etudiants.filter(decision_generale=decision_filter)

    # Filtre par recherche
    if search_query:
        etudiants = etudiants.filter(
            Q(nom_prenom__icontains=search_query) |
            Q(matricule__icontains=search_query)
        )

    # Filtre par moyenne min/max
    if moy_min:
        try:
            etudiants = etudiants.filter(moyenne_generale__gte=Decimal(moy_min))
        except:
            pass

    if moy_max:
        try:
            etudiants = etudiants.filter(moyenne_generale__lte=Decimal(moy_max))
        except:
            pass

    # Filtre par ECUE spécifique
    ecue_obj = None
    if ecue_filter:
        try:
            ecue_obj = ECUE.objects.get(code=ecue_filter, ue__pv=pv)
        except ECUE.DoesNotExist:
            pass

        if decision_ecue_filter:
            etudiants = etudiants.filter(
                notes__ecue__code=ecue_filter,
                notes__decision=decision_ecue_filter
            ).distinct()
        else:
            etudiants = etudiants.filter(notes__ecue__code=ecue_filter).distinct()

    # Tri par nom (optionnel mais recommandé pour émargement)
    etudiants = etudiants.order_by('nom_prenom')

    # Créer le fichier Excel simplifié
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille Émargement"

    # Styles
    header_font = Font(bold=True, size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # En-tête du document
    ws['A1'] = 'UNIVERSITÉ DE DOUALA'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')
    ws['A1'].alignment = center_alignment

    ws['A2'] = 'École Nationale Supérieure Polytechnique de Douala'
    ws['A2'].font = Font(bold=True, size=12)
    ws.merge_cells('A2:D2')
    ws['A2'].alignment = center_alignment

    ws['A3'] = f"FEUILLE D'ÉMARGEMENT - {pv.filiere} - {pv.niveau} - {pv.semestre}"
    ws['A3'].font = Font(bold=True, size=11)
    ws.merge_cells('A3:D3')
    ws['A3'].alignment = center_alignment

    ws['A4'] = f"Année académique: {pv.annee_academique}"
    ws['A4'].font = Font(size=10)
    ws.merge_cells('A4:D4')
    ws['A4'].alignment = center_alignment

    # Ligne 5 : Matière filtrée (si applicable)
    current_row = 5
    if ecue_obj:
        ws[f'A{current_row}'] = f"Matière : {ecue_obj.code} - {ecue_obj.intitule}"
        ws[f'A{current_row}'].font = Font(size=10, bold=True)
        ws.merge_cells(f'A{current_row}:D{current_row}')
        ws[f'A{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
        current_row += 1

    # Ligne vide
    current_row += 1

    # En-têtes des colonnes
    headers = ['N°', 'MATRICULE', 'NOM & PRÉNOMS', 'SIGNATURE']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Données des étudiants
    current_row += 1
    for idx, etudiant in enumerate(etudiants, start=1):
        # N°
        cell = ws.cell(row=current_row, column=1, value=idx)
        cell.border = border
        cell.alignment = center_alignment

        # Matricule
        cell = ws.cell(row=current_row, column=2, value=etudiant.matricule)
        cell.border = border
        cell.alignment = center_alignment

        # Nom & Prénoms
        cell = ws.cell(row=current_row, column=3, value=etudiant.nom_prenom)
        cell.border = border
        cell.alignment = Alignment(vertical="center")

        # Signature (vide)
        cell = ws.cell(row=current_row, column=4, value="")
        cell.border = border

        # Hauteur de ligne augmentée pour faciliter la signature manuscrite
        ws.row_dimensions[current_row].height = 30

        current_row += 1

    # Largeurs de colonnes optimisées
    ws.column_dimensions['A'].width = 6    # N°
    ws.column_dimensions['B'].width = 18   # Matricule
    ws.column_dimensions['C'].width = 40   # Nom & Prénoms
    ws.column_dimensions['D'].width = 30   # Signature

    # Nom du fichier avec date
    date_str = datetime.now().strftime('%Y-%m-%d')
    niveau_str = str(pv.niveau).replace('/', '-') if pv.niveau else 'Niveau'
    filename = f"Feuille_Emargement_{niveau_str}_{date_str}.xlsx"

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def dashboard_aggrid(request, pk):
    """
    Redirection vers le dashboard principal (les deux vues ont été fusionnées)
    """
    # Rediriger vers le dashboard principal en conservant les paramètres GET
    url = reverse('pv:dashboard', args=[pk])
    if request.GET:
        url += '?' + request.GET.urlencode()
    return redirect(url)


def export_emargements_nv_complets(request, pk):
    """
    NOUVELLE FONCTIONNALITÉ : Exporte un fichier Excel multi-feuilles
    avec les émargements NV par matière (ECUE).

    Chaque feuille contient:
    - Les étudiants ayant obtenu NV dans cette matière
    - Leurs notes (CC, EX, Moyenne)
    - Une colonne pour signature manuscrite
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # Compteur de feuilles créées
    feuilles_creees = 0

    # Styles communs
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Parcourir toutes les UE du PV
    for ue in pv.ues.all().prefetch_related('ecues').order_by('ordre'):
        # Parcourir tous les ECUE de l'UE
        for ecue in ue.ecues.all().order_by('ordre'):

            # Récupérer les étudiants NV pour cet ECUE
            notes_nv = Note.objects.filter(
                ecue=ecue,
                decision='NV'
            ).select_related('etudiant').order_by('etudiant__nom_prenom')

            # Si aucun étudiant NV, passer à l'ECUE suivant
            if not notes_nv.exists():
                continue

            # Créer une feuille pour cet ECUE
            # Nom de feuille limité à 31 caractères (limite Excel)
            nom_feuille = ecue.code[:31] if len(ecue.code) <= 31 else ecue.code[:28] + "..."
            ws = wb.create_sheet(title=nom_feuille)

            # ===== EN-TÊTE DE LA FEUILLE =====

            # Ligne 2 : Nom de l'école
            ws.merge_cells('A2:D2')
            cell_ecole = ws['A2']
            cell_ecole.value = "ÉCOLE NATIONALE SUPÉRIEURE POLYTECHNIQUE DE DOUALA"
            cell_ecole.font = Font(size=14, bold=True)
            cell_ecole.alignment = Alignment(horizontal='center', vertical='center')

            # Ligne 3 : Titre
            ws.merge_cells('A3:D3')
            cell_titre = ws['A3']
            cell_titre.value = "FEUILLE D'ÉMARGEMENT - ÉTUDIANTS NON VALIDÉS"
            cell_titre.font = Font(size=12, bold=True)
            cell_titre.alignment = Alignment(horizontal='center', vertical='center')

            # Ligne 5 : Matière
            ws.merge_cells('A5:D5')
            cell_matiere = ws['A5']
            cell_matiere.value = f"Matière : {ecue.code} - {ecue.intitule}"
            cell_matiere.font = Font(size=11, bold=True)
            cell_matiere.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 6 : UE
            ws.merge_cells('A6:D6')
            cell_ue = ws['A6']
            cell_ue.value = f"UE : {ue.code} - {ue.intitule}"
            cell_ue.font = Font(size=10)
            cell_ue.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 7 : Niveau et Semestre
            ws.merge_cells('A7:D7')
            cell_niveau = ws['A7']
            cell_niveau.value = f"Niveau : {pv.filiere} {pv.niveau} | Semestre : {pv.semestre}"
            cell_niveau.font = Font(size=10)
            cell_niveau.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 8 : Année académique
            ws.merge_cells('A8:D8')
            cell_annee = ws['A8']
            cell_annee.value = f"Année académique : {pv.annee_academique}"
            cell_annee.font = Font(size=10)
            cell_annee.alignment = Alignment(horizontal='left', vertical='center')

            # ===== EN-TÊTES DU TABLEAU (Ligne 10) =====

            headers = ['N°', 'MATRICULE', 'NOM & PRÉNOMS', 'SIGNATURE']

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=10, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # ===== DONNÉES DES ÉTUDIANTS NV =====

            current_row = 11

            for idx, note in enumerate(notes_nv, start=1):
                etudiant = note.etudiant

                # Colonne 1 : N°
                cell_num = ws.cell(row=current_row, column=1)
                cell_num.value = idx
                cell_num.alignment = Alignment(horizontal='center', vertical='center')
                cell_num.border = border

                # Colonne 2 : Matricule
                cell_mat = ws.cell(row=current_row, column=2)
                cell_mat.value = etudiant.matricule
                cell_mat.alignment = Alignment(horizontal='center', vertical='center')
                cell_mat.border = border

                # Colonne 3 : Nom & Prénoms
                cell_nom = ws.cell(row=current_row, column=3)
                cell_nom.value = etudiant.nom_prenom
                cell_nom.alignment = Alignment(horizontal='left', vertical='center')
                cell_nom.border = border

                # Colonne 4 : SIGNATURE (vide)
                cell_sig = ws.cell(row=current_row, column=4)
                cell_sig.value = ""
                cell_sig.border = border

                # Hauteur de ligne pour signature manuscrite
                ws.row_dimensions[current_row].height = 30

                current_row += 1

            # ===== PIED DE PAGE =====

            # Ligne N+2 : Total étudiants NV
            total_row = current_row + 2
            ws.merge_cells(f'A{total_row}:D{total_row}')
            cell_total = ws[f'A{total_row}']
            cell_total.value = f"Total étudiants NV pour cette matière : {notes_nv.count()}"
            cell_total.font = Font(bold=True, size=11)
            cell_total.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne N+4 : Date et Signature enseignant
            signature_row = total_row + 2
            ws.merge_cells(f'A{signature_row}:D{signature_row}')
            cell_date = ws[f'A{signature_row}']
            cell_date.value = "Date : _______________    Signature enseignant : _______________"
            cell_date.alignment = Alignment(horizontal='left', vertical='center')

            # ===== LARGEURS DES COLONNES =====

            ws.column_dimensions['A'].width = 6   # N°
            ws.column_dimensions['B'].width = 18  # Matricule
            ws.column_dimensions['C'].width = 40  # Nom & Prénoms
            ws.column_dimensions['D'].width = 30  # Signature

            feuilles_creees += 1

    # Vérifier qu'au moins une feuille a été créée
    if feuilles_creees == 0:
        # Aucun étudiant NV trouvé
        ws = wb.create_sheet(title="Information")
        ws['A1'] = "Aucun étudiant Non Validé (NV) trouvé dans ce PV."
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30

    # Nom du fichier
    date_str = datetime.now().strftime('%Y-%m-%d')
    filiere_clean = pv.filiere.replace('/', '-').replace('\\', '-')[:20]
    filename = f"Emargements_NV_{filiere_clean}_{pv.niveau}_{pv.semestre}_{date_str}.xlsx"

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def export_emargements_v_vc(request, pk):
    """
    NOUVELLE FONCTIONNALITÉ : Exporte un fichier Excel multi-feuilles
    avec les émargements V et VC par matière (ECUE).

    Chaque feuille contient:
    - Les étudiants ayant obtenu V ou VC dans cette matière
    - Format simple pour impression et signatures
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Créer le workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Supprimer la feuille par défaut

    # Compteur de feuilles créées
    feuilles_creees = 0

    # Styles communs
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Parcourir toutes les UE du PV
    for ue in pv.ues.all().prefetch_related('ecues').order_by('ordre'):
        # Parcourir tous les ECUE de l'UE
        for ecue in ue.ecues.all().order_by('ordre'):

            # Récupérer les étudiants V ou VC pour cet ECUE
            notes_v_vc = Note.objects.filter(
                ecue=ecue,
                decision__in=['V', 'VC']
            ).select_related('etudiant').order_by('etudiant__nom_prenom')

            # Si aucun étudiant V ou VC, passer à l'ECUE suivant
            if not notes_v_vc.exists():
                continue

            # Créer une feuille pour cet ECUE
            # Nom de feuille limité à 31 caractères (limite Excel)
            nom_feuille = ecue.code[:31] if len(ecue.code) <= 31 else ecue.code[:28] + "..."
            ws = wb.create_sheet(title=nom_feuille)

            # ===== EN-TÊTE DE LA FEUILLE =====

            # Ligne 2 : Nom de l'école
            ws.merge_cells('A2:D2')
            cell_ecole = ws['A2']
            cell_ecole.value = "ÉCOLE NATIONALE SUPÉRIEURE POLYTECHNIQUE DE DOUALA"
            cell_ecole.font = Font(size=14, bold=True)
            cell_ecole.alignment = Alignment(horizontal='center', vertical='center')

            # Ligne 3 : Titre
            ws.merge_cells('A3:D3')
            cell_titre = ws['A3']
            cell_titre.value = "FEUILLE D'ÉMARGEMENT - ÉTUDIANTS VALIDÉS (V et VC)"
            cell_titre.font = Font(size=12, bold=True)
            cell_titre.alignment = Alignment(horizontal='center', vertical='center')

            # Ligne 5 : Matière
            ws.merge_cells('A5:D5')
            cell_matiere = ws['A5']
            cell_matiere.value = f"Matière : {ecue.code} - {ecue.intitule}"
            cell_matiere.font = Font(size=11, bold=True)
            cell_matiere.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 6 : UE
            ws.merge_cells('A6:D6')
            cell_ue = ws['A6']
            cell_ue.value = f"UE : {ue.code} - {ue.intitule}"
            cell_ue.font = Font(size=10)
            cell_ue.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 7 : Niveau et Semestre
            ws.merge_cells('A7:D7')
            cell_niveau = ws['A7']
            cell_niveau.value = f"Niveau : {pv.filiere} {pv.niveau} | Semestre : {pv.semestre}"
            cell_niveau.font = Font(size=10)
            cell_niveau.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne 8 : Année académique
            ws.merge_cells('A8:D8')
            cell_annee = ws['A8']
            cell_annee.value = f"Année académique : {pv.annee_academique}"
            cell_annee.font = Font(size=10)
            cell_annee.alignment = Alignment(horizontal='left', vertical='center')

            # ===== EN-TÊTES DU TABLEAU (Ligne 10) =====

            headers = ['N°', 'MATRICULE', 'NOM & PRÉNOMS', 'SIGNATURE']

            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=10, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # ===== DONNÉES DES ÉTUDIANTS V et VC =====

            current_row = 11

            for idx, note in enumerate(notes_v_vc, start=1):
                etudiant = note.etudiant

                # Colonne 1 : N°
                cell_num = ws.cell(row=current_row, column=1)
                cell_num.value = idx
                cell_num.alignment = Alignment(horizontal='center', vertical='center')
                cell_num.border = border

                # Colonne 2 : Matricule
                cell_mat = ws.cell(row=current_row, column=2)
                cell_mat.value = etudiant.matricule
                cell_mat.alignment = Alignment(horizontal='center', vertical='center')
                cell_mat.border = border

                # Colonne 3 : Nom & Prénoms
                cell_nom = ws.cell(row=current_row, column=3)
                cell_nom.value = etudiant.nom_prenom
                cell_nom.alignment = Alignment(horizontal='left', vertical='center')
                cell_nom.border = border

                # Colonne 4 : SIGNATURE (vide)
                cell_sig = ws.cell(row=current_row, column=4)
                cell_sig.value = ""
                cell_sig.border = border

                # Hauteur de ligne pour signature manuscrite
                ws.row_dimensions[current_row].height = 30

                current_row += 1

            # ===== PIED DE PAGE =====

            # Ligne N+2 : Total étudiants V et VC
            total_row = current_row + 2
            ws.merge_cells(f'A{total_row}:D{total_row}')
            cell_total = ws[f'A{total_row}']
            cell_total.value = f"Total étudiants validés (V et VC) pour cette matière : {notes_v_vc.count()}"
            cell_total.font = Font(bold=True, size=11)
            cell_total.alignment = Alignment(horizontal='left', vertical='center')

            # Ligne N+4 : Date et Signature enseignant
            signature_row = total_row + 2
            ws.merge_cells(f'A{signature_row}:D{signature_row}')
            cell_date = ws[f'A{signature_row}']
            cell_date.value = "Date : _______________    Signature enseignant : _______________"
            cell_date.alignment = Alignment(horizontal='left', vertical='center')

            # ===== LARGEURS DES COLONNES =====

            ws.column_dimensions['A'].width = 6   # N°
            ws.column_dimensions['B'].width = 18  # Matricule
            ws.column_dimensions['C'].width = 40  # Nom & Prénoms
            ws.column_dimensions['D'].width = 30  # Signature

            feuilles_creees += 1

    # Vérifier qu'au moins une feuille a été créée
    if feuilles_creees == 0:
        # Aucun étudiant V ou VC trouvé
        ws = wb.create_sheet(title="Information")
        ws['A1'] = "Aucun étudiant Validé (V ou VC) trouvé dans ce PV."
        ws['A1'].font = Font(size=12, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:E1')
        ws.row_dimensions[1].height = 30

    # Nom du fichier
    date_str = datetime.now().strftime('%Y-%m-%d')
    filiere_clean = pv.filiere.replace('/', '-').replace('\\', '-')[:20]
    filename = f"Emargements_V_VC_{filiere_clean}_{pv.niveau}_{pv.semestre}_{date_str}.xlsx"

    # Réponse HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)

    return response


def print_view(request, pk):
    """
    Vue optimisée pour l'impression
    """
    pv = get_object_or_404(ProcesVerbal, pk=pk)

    # Appliquer les mêmes filtres
    decision_filter = request.GET.get('decision', '')
    etudiants = pv.etudiants.all()

    if decision_filter:
        etudiants = etudiants.filter(decision_generale=decision_filter)

    context = {
        'pv': pv,
        'etudiants': etudiants,
    }

    return render(request, 'pv/print.html', context)
