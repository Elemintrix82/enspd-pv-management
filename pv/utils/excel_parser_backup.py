"""
Utilitaire pour parser les fichiers Excel de PV de délibération
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from decimal import Decimal
import re


class PVExcelParser:
    """
    Classe pour parser et extraire les données d'un fichier Excel PV
    """

    def __init__(self, file_path):
        """
        Initialise le parser avec le chemin du fichier Excel

        Args:
            file_path: Chemin vers le fichier Excel
        """
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.metadata = {}
        self.ues = []
        self.ecues = []
        self.etudiants = []
        self.header_row = None

    def parse(self):
        """
        Parse le fichier Excel et extrait toutes les données

        Returns:
            dict: Dictionnaire contenant toutes les données extraites
        """
        self.extract_metadata()
        self.extract_ue_structure()
        self.extract_ecue_structure()
        self.find_header_row()
        self.extract_student_data()

        return {
            'metadata': self.metadata,
            'ues': self.ues,
            'ecues': self.ecues,
            'etudiants': self.etudiants
        }

    def extract_metadata(self):
        """
        Extrait les métadonnées institutionnelles du fichier Excel
        (lignes 1-7 généralement)
        """
        # Université (ligne 1, colonne F)
        self.metadata['universite'] = self.ws['F1'].value or "UNIVERSITE DE DOUALA"

        # École (ligne 3, colonne F)
        self.metadata['ecole'] = self.ws['F3'].value or "École Nationale Supérieure Polytechnique de Douala"

        # Niveau (ligne 4, colonne I)
        niveau_val = self.ws['I4'].value
        self.metadata['niveau'] = int(niveau_val) if niveau_val else 4

        # Filière (ligne 8, colonne F)
        filiere_val = self.ws['F8'].value
        if filiere_val and ':' in str(filiere_val):
            self.metadata['filiere'] = str(filiere_val).split(':', 1)[1].strip()
        else:
            self.metadata['filiere'] = str(filiere_val) if filiere_val else "GRT"

        # Semestre (ligne 7, colonnes H et I)
        semestre_s7 = self.ws['H7'].value
        self.metadata['semestre'] = str(semestre_s7) if semestre_s7 else "S7"

        # Année académique (chercher dans les lignes 5-6)
        annee_found = False
        for row in range(5, 7):
            for col in range(6, 12):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and '/' in str(cell_val) and len(str(cell_val).strip()) <= 12:
                    # Format probable: 2022/2023
                    self.metadata['annee_academique'] = str(cell_val).strip()
                    annee_found = True
                    break
            if annee_found:
                break

        if not annee_found:
            self.metadata['annee_academique'] = "2022/2023"

        # Formation (chercher le mot "ALTERNANCE" ou "CLASSIQUE")
        formation_found = False
        for row in range(1, 10):
            for col in range(1, 15):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and 'ALTERNANCE' in str(cell_val).upper():
                    self.metadata['formation'] = "ALTERNANCE"
                    formation_found = True
                    break
                elif cell_val and 'CLASSIQUE' in str(cell_val).upper():
                    self.metadata['formation'] = "CLASSIQUE"
                    formation_found = True
                    break
            if formation_found:
                break

        if not formation_found:
            self.metadata['formation'] = "ALTERNANCE"

    def extract_ue_structure(self):
        """
        Extrait la structure des UE (ligne 9, généralement)
        """
        ue_row = 9  # Ligne 9 (index 8 en 0-based)
        ue_ordre = 1

        for col_idx in range(5, self.ws.max_column + 1):
            cell_val = self.ws.cell(row=ue_row, column=col_idx).value

            if cell_val and ('EPDGIT' in str(cell_val) or 'EPDTCO' in str(cell_val)):
                # Extraire code et intitulé
                parts = str(cell_val).split(':', 1)
                if len(parts) == 2:
                    ue_code = parts[0].strip()
                    ue_intitule = parts[1].strip()

                    # Vérifier si cette UE n'existe pas déjà
                    if not any(ue['code'] == ue_code for ue in self.ues):
                        self.ues.append({
                            'code': ue_code,
                            'intitule': ue_intitule,
                            'ordre': ue_ordre,
                            'colonne_debut': col_idx
                        })
                        ue_ordre += 1

    def extract_ecue_structure(self):
        """
        Extrait la structure des ECUE (ligne 10, généralement)
        """
        ecue_row = 10  # Ligne 10 (index 9 en 0-based)
        ecue_ordre = 1

        for col_idx in range(5, self.ws.max_column + 1):
            cell_val = self.ws.cell(row=ecue_row, column=col_idx).value

            if cell_val and ('EPDGIT' in str(cell_val) or 'EPDTCO' in str(cell_val)):
                # Extraire code et intitulé
                if ')' in str(cell_val):
                    code_part = str(cell_val).split(')')[0]
                    ecue_code = code_part.replace('(', '').strip()
                    ecue_intitule = str(cell_val).split(')', 1)[1].strip() if ')' in str(cell_val) else ''

                    # Déterminer l'UE parente
                    ue_parent_code = None
                    ecue_code_prefix = ecue_code[:9]  # Ex: EPDGIT415

                    for ue in self.ues:
                        if ue['code'] == ecue_code_prefix:
                            ue_parent_code = ue['code']
                            break

                    # Vérifier si cette ECUE n'existe pas déjà
                    if not any(ecue['code'] == ecue_code for ecue in self.ecues):
                        self.ecues.append({
                            'code': ecue_code,
                            'intitule': ecue_intitule,
                            'ordre': ecue_ordre,
                            'ue_code': ue_parent_code,
                            'colonne_debut': col_idx
                        })
                        ecue_ordre += 1

    def find_header_row(self):
        """
        Trouve la ligne d'en-tête contenant N°, MATRICULE, NOMS & PRENOMS
        """
        for row_idx in range(1, 20):
            for col_idx in range(1, 10):
                cell_val = self.ws.cell(row=row_idx, column=col_idx).value
                if cell_val and 'N°' in str(cell_val) or 'MATRICULE' in str(cell_val).upper():
                    self.header_row = row_idx
                    return

        # Par défaut, ligne 11 (index 10 en 0-based)
        self.header_row = 11

    def extract_student_data(self):
        """
        Extrait les données des étudiants
        """
        # Lire le fichier avec pandas à partir de la ligne d'en-tête
        df = pd.read_excel(self.file_path, header=self.header_row - 1)

        # Nettoyer les noms de colonnes
        df.columns = [str(col).strip() if pd.notna(col) else f"Unnamed_{i}"
                      for i, col in enumerate(df.columns)]

        for idx, row in df.iterrows():
            # Extraire les informations d'identité
            numero = row.get('N°', idx + 1)
            matricule = str(row.get('MATRICULE', '')).strip()
            nom_prenom = str(row.get('NOMS & PRENOMS', '')).strip()

            # Si matricule vide, ignorer cette ligne
            if not matricule or matricule == 'nan':
                continue

            # Extraire la moyenne générale et décision
            moyenne_generale = self._safe_decimal(row.get('MOYENNE/20', 0))
            credits_acquis = self._safe_int(row.get('CREDITS  ACQUIS', 0))

            # Décision générale (dernière colonne DECISION)
            decision_generale = self._extract_decision(row, df.columns[-1])

            # Extraire les notes par ECUE
            notes_par_ecue = []
            for ecue in self.ecues:
                note_data = self._extract_note_for_ecue(row, df.columns, ecue)
                if note_data:
                    notes_par_ecue.append(note_data)

            etudiant_data = {
                'numero': self._safe_int(numero),
                'matricule': matricule,
                'nom_prenom': nom_prenom,
                'moyenne_generale': moyenne_generale,
                'credits_acquis': credits_acquis,
                'decision_generale': decision_generale,
                'notes': notes_par_ecue
            }

            self.etudiants.append(etudiant_data)

    def _extract_note_for_ecue(self, row, columns, ecue):
        """
        Extrait les notes d'un étudiant pour une ECUE donnée

        Args:
            row: Ligne de données pandas
            columns: Liste des colonnes
            ecue: Dictionnaire ECUE

        Returns:
            dict: Données de note ou None
        """
        # Chercher les colonnes CC, EX, MOY, CA, DECISION pour cette ECUE
        # On se base sur l'ordre des colonnes

        note_data = {
            'ecue_code': ecue['code'],
            'cc': None,
            'examen': None,
            'moyenne': None,
            'credit_attribue': 0,
            'decision': 'NV'
        }

        # Stratégie: chercher dans les colonnes par pattern
        # On va essayer de trouver la séquence CC, EX, MOY, CA, DECISION

        for i, col in enumerate(columns):
            col_str = str(col).upper()

            # Détecter les colonnes de notes
            if 'CC' in col_str and i + 4 < len(columns):
                # Vérifier si c'est la bonne séquence
                potential_cc = self._safe_decimal(row.get(columns[i]))
                potential_ex = self._safe_decimal(row.get(columns[i + 1]))
                potential_moy = self._safe_decimal(row.get(columns[i + 2]))
                potential_ca = self._safe_int(row.get(columns[i + 3]))
                potential_decision = self._extract_decision(row, columns[i + 4])

                # Si on a une moyenne valide, c'est probablement une note
                if potential_moy and potential_moy > 0:
                    note_data['cc'] = potential_cc
                    note_data['examen'] = potential_ex
                    note_data['moyenne'] = potential_moy
                    note_data['credit_attribue'] = potential_ca
                    note_data['decision'] = potential_decision
                    return note_data

        # Méthode alternative: on retourne une note par défaut
        return note_data if note_data['moyenne'] else None

    def _safe_decimal(self, value):
        """Convertit une valeur en Decimal de manière sécurisée"""
        try:
            if pd.isna(value):
                return Decimal('0.00')
            return Decimal(str(value))
        except:
            return Decimal('0.00')

    def _safe_int(self, value):
        """Convertit une valeur en int de manière sécurisée"""
        try:
            if pd.isna(value):
                return 0
            return int(float(value))
        except:
            return 0

    def _extract_decision(self, row, column_name):
        """
        Extrait la décision (V, NV, VC) d'une colonne

        Args:
            row: Ligne pandas
            column_name: Nom de la colonne

        Returns:
            str: Décision (V, NV, VC) ou 'NV' par défaut
        """
        try:
            decision_val = str(row.get(column_name, 'NV')).strip().upper()

            # IMPORTANT: Vérifier "NON VALIDE" AVANT "VALIDE"
            if 'NON VALIDE' in decision_val or decision_val == 'NV':
                return 'NV'
            elif 'COMPENSATION' in decision_val or decision_val == 'VC':
                return 'VC'
            elif 'VALIDE' in decision_val or decision_val == 'V':
                return 'V'
            else:
                # Par défaut
                return 'NV'
        except:
            return 'NV'
