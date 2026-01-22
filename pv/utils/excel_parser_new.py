"""
Nouveau parser Excel optimisé pour extraire les notes ECUE et synthèses UE
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from decimal import Decimal
import re


class PVExcelParser:
    """
    Parser amélioré pour extraire les données PV avec gestion correcte des synthèses UE
    """

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.metadata = {}
        self.ues = []
        self.ecues = []
        self.etudiants = []
        self.header_row = 11  # Ligne 11 (index 10 en 0-based)

    def parse(self):
        """Parse complet du fichier Excel"""
        self.extract_metadata()
        self.extract_structure()
        self.extract_student_data()

        return {
            'metadata': self.metadata,
            'ues': self.ues,
            'ecues': self.ecues,
            'etudiants': self.etudiants
        }

    def extract_metadata(self):
        """Extrait les métadonnées du fichier"""
        # Université (ligne 1, colonne F)
        self.metadata['universite'] = self.ws['F1'].value or "UNIVERSITE DE DOUALA"

        # École (ligne 3, colonne F)
        self.metadata['ecole'] = self.ws['F3'].value or "École Nationale Supérieure Polytechnique de Douala"

        # Niveau (ligne 4, colonne I)
        niveau_val = self.ws['I4'].value
        self.metadata['niveau'] = int(niveau_val) if niveau_val else 4

        # Filière (ligne 8, colonne F)
        filiere_val = self.ws['F8'].value
        if filiere_val and ':' in str(filiere_val):
            self.metadata['filiere'] = str(filiere_val).split(':', 1)[1].strip()
        else:
            self.metadata['filiere'] = str(filiere_val) if filiere_val else "GRT"

        # Semestre (ligne 7, colonnes H et I)
        semestre_s7 = self.ws['H7'].value
        self.metadata['semestre'] = str(semestre_s7) if semestre_s7 else "S7"

        # Année académique (chercher dans les lignes 5-6)
        annee_found = False
        for row in range(5, 7):
            for col in range(6, 12):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and '/' in str(cell_val) and len(str(cell_val).strip()) <= 12:
                    self.metadata['annee_academique'] = str(cell_val).strip()
                    annee_found = True
                    break
            if annee_found:
                break

        if not annee_found:
            self.metadata['annee_academique'] = "2022/2023"

        # Formation (chercher le mot "ALTERNANCE" ou "CLASSIQUE")
        formation_found = False
        for row in range(1, 10):
            for col in range(1, 15):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and 'ALTERNANCE' in str(cell_val).upper():
                    self.metadata['formation'] = "ALTERNANCE"
                    formation_found = True
                    break
                elif cell_val and 'CLASSIQUE' in str(cell_val).upper():
                    self.metadata['formation'] = "CLASSIQUE"
                    formation_found = True
                    break
            if formation_found:
                break

        if not formation_found:
            self.metadata['formation'] = "ALTERNANCE"

    def extract_structure(self):
        """
        Extrait la structure UE/ECUE en analysant les lignes 9, 10, 11
        """
        df_structure = pd.read_excel(self.file_path, header=None, nrows=11)

        ue_row = df_structure.iloc[8]  # Ligne 9
        ecue_row = df_structure.iloc[9]  # Ligne 10
        header_row = df_structure.iloc[10]  # Ligne 11

        ue_ordre = 1
        ecue_ordre = 1
        current_ue = None
        current_ue_code = None

        # Mapper les colonnes
        col_idx = 0
        while col_idx < len(header_row):
            ue_val = ue_row.iloc[col_idx] if col_idx < len(ue_row) else None
            ecue_val = ecue_row.iloc[col_idx] if col_idx < len(ecue_row) else None
            header_val = header_row.iloc[col_idx] if col_idx < len(header_row) else None

            # Détection d'une nouvelle UE
            if pd.notna(ue_val) and ('EPDGIT' in str(ue_val) or 'EPDTCO' in str(ue_val)):
                parts = str(ue_val).split(':', 1)
                if len(parts) == 2:
                    ue_code = parts[0].strip()
                    ue_intitule = parts[1].strip()

                    # Vérifier si nouvelle UE
                    if not any(ue['code'] == ue_code for ue in self.ues):
                        current_ue = {
                            'code': ue_code,
                            'intitule': ue_intitule,
                            'ordre': ue_ordre,
                            'col_start': col_idx
                        }
                        self.ues.append(current_ue)
                        current_ue_code = ue_code
                        ue_ordre += 1

            # Détection d'une ECUE
            if pd.notna(ecue_val):
                ecue_str = str(ecue_val).strip()

                # Cas 1: ECUE normale (EPDGIT4151)
                if '(' in ecue_str and ')' in ecue_str and ('EPDGIT' in ecue_str or 'EPDTCO' in ecue_str):
                    code_part = ecue_str.split(')')[0]
                    ecue_code = code_part.replace('(', '').strip()
                    ecue_intitule = ecue_str.split(')', 1)[1].strip() if ')' in ecue_str else ''

                    # Déterminer l'UE parente
                    ue_parent_code = current_ue_code

                    # Vérifier si nouvelle ECUE
                    if not any(e['code'] == ecue_code for e in self.ecues):
                        self.ecues.append({
                            'code': ecue_code,
                            'intitule': ecue_intitule,
                            'ordre': ecue_ordre,
                            'ue_code': ue_parent_code,
                            'col_start': col_idx,
                            'is_synthese': False
                        })
                        ecue_ordre += 1

                # Cas 2: SYNTHESE UE
                elif 'SYNTHESE' in ecue_str.upper() and current_ue_code:
                    # C'est une synthèse UE, on l'enregistre différemment
                    synthese_code = f"SYNTHESE_{current_ue_code}"

                    if not any(e['code'] == synthese_code for e in self.ecues):
                        self.ecues.append({
                            'code': synthese_code,
                            'intitule': f"Synthèse {current_ue_code}",
                            'ordre': ecue_ordre,
                            'ue_code': current_ue_code,
                            'col_start': col_idx,
                            'is_synthese': True  # IMPORTANT
                        })
                        ecue_ordre += 1

            col_idx += 1

    def extract_student_data(self):
        """Extrait les données des étudiants avec notes et synthèses"""
        # Lire avec pandas
        df = pd.read_excel(self.file_path, header=self.header_row - 1)

        # Nettoyer les colonnes
        df.columns = [str(col).strip() if pd.notna(col) else f"Unnamed_{i}"
                      for i, col in enumerate(df.columns)]

        for idx, row in df.iterrows():
            # Informations de base
            numero = row.get('N°', idx + 1)
            matricule = str(row.get('MATRICULE', '')).strip()
            nom_prenom = str(row.get('NOMS & PRENOMS', '')).strip()

            # Si matricule vide, ignorer
            if not matricule or matricule == 'nan':
                continue

            # Moyenne générale et crédits
            moyenne_generale = self._safe_decimal(row.get('MOYENNE/20', 0))
            credits_acquis = self._safe_int(row.get('CREDITS  ACQUIS', 0))

            # Décision générale (dernière colonne DECISION)
            decision_generale = self._extract_decision(row, df.columns[-1])

            # Extraire les notes par ECUE et synthèses UE
            notes_par_ecue = []
            syntheses_par_ue = []

            for ecue in self.ecues:
                if ecue.get('is_synthese', False):
                    # C'est une synthèse UE (seulement MOY, CA, DEC)
                    synthese_data = self._extract_synthese_ue(row, df.columns, ecue)
                    if synthese_data:
                        syntheses_par_ue.append(synthese_data)
                else:
                    # C'est une ECUE normale (CC, EX, MOY, CA, DEC)
                    note_data = self._extract_note_for_ecue(row, df.columns, ecue)
                    if note_data:
                        notes_par_ecue.append(note_data)

            etudiant_data = {
                'numero': self._safe_int(numero),
                'matricule': matricule,
                'nom_prenom': nom_prenom,
                'moyenne_generale': moyenne_generale,
                'credits_acquis': credits_acquis,
                'decision_generale': decision_generale,
                'notes': notes_par_ecue,
                'syntheses_ue': syntheses_par_ue
            }

            self.etudiants.append(etudiant_data)

    def _extract_note_for_ecue(self, row, columns, ecue):
        """
        Extrait une note ECUE avec CC, EX, MOY, CA, DECISION
        """
        note_data = {
            'ecue_code': ecue['code'],
            'cc': None,
            'examen': None,
            'moyenne': None,
            'credit_attribue': 0,
            'decision': 'NV'
        }

        # Chercher les colonnes CC, EX, MOY, CA, DECISION
        # Stratégie: parcourir les colonnes et chercher la séquence
        for i, col in enumerate(columns):
            col_str = str(col).upper()

            # Détecter CC suivi de EX, MOY, CA, DECISION
            if ('CC' in col_str or col_str == 'CC') and i + 4 < len(columns):
                # Vérifier la séquence
                next_1 = str(columns[i + 1]).upper()
                next_2 = str(columns[i + 2]).upper()
                next_3 = str(columns[i + 3]).upper()
                next_4 = str(columns[i + 4]).upper()

                # Séquence attendue: CC, EX, MOY, CA, DECISION
                if ('EX' in next_1) and ('MOY' in next_2):
                    potential_cc = self._safe_decimal(row.get(columns[i]))
                    potential_ex = self._safe_decimal(row.get(columns[i + 1]))
                    potential_moy = self._safe_decimal(row.get(columns[i + 2]))
                    potential_ca = self._safe_int(row.get(columns[i + 3]))
                    potential_decision = self._extract_decision(row, columns[i + 4])

                    # Si moyenne > 0, c'est probablement une note valide
                    if potential_moy and potential_moy > 0:
                        note_data['cc'] = potential_cc
                        note_data['examen'] = potential_ex
                        note_data['moyenne'] = potential_moy
                        note_data['credit_attribue'] = potential_ca
                        note_data['decision'] = potential_decision
                        return note_data

        return note_data if note_data['moyenne'] else None

    def _extract_synthese_ue(self, row, columns, ecue):
        """
        Extrait une synthèse UE avec MOY, CA, DECISION (pas de CC/EX)
        """
        synthese_data = {
            'ue_code': ecue['ue_code'],
            'moyenne_ue': None,
            'credits_attribues': 0,
            'decision': 'NV'
        }

        # Chercher MOY, CA, DECISION pour synthèse
        # Stratégie: chercher "SYNTHESE" ou MOY isolée
        for i, col in enumerate(columns):
            col_str = str(col).upper()

            # Détecter MOY suivi de CA, DECISION (sans CC/EX avant)
            if i > 0 and i + 2 < len(columns):
                prev_col = str(columns[i - 1]).upper()
                next_1 = str(columns[i + 1]).upper()
                next_2 = str(columns[i + 2]).upper()

                # Si c'est MOY sans CC/EX juste avant
                if ('MOY' in col_str) and ('CC' not in prev_col) and ('EX' not in prev_col):
                    # Et si suivi de CA, DECISION
                    if ('CA' in next_1 or 'UNNAMED' in next_1) and ('DECISION' in next_2 or 'DEC' in next_2):
                        potential_moy = self._safe_decimal(row.get(columns[i]))
                        potential_ca = self._safe_int(row.get(columns[i + 1]))
                        potential_decision = self._extract_decision(row, columns[i + 2])

                        if potential_moy and potential_moy > 0:
                            synthese_data['moyenne_ue'] = potential_moy
                            synthese_data['credits_attribues'] = potential_ca
                            synthese_data['decision'] = potential_decision
                            return synthese_data

        return synthese_data if synthese_data['moyenne_ue'] else None

    def _safe_decimal(self, value):
        """Convertit en Decimal de manière sécurisée"""
        try:
            if pd.isna(value):
                return Decimal('0.00')
            return Decimal(str(value))
        except:
            return Decimal('0.00')

    def _safe_int(self, value):
        """Convertit en int de manière sécurisée"""
        try:
            if pd.isna(value):
                return 0
            return int(float(value))
        except:
            return 0

    def _extract_decision(self, row, column_name):
        """Extrait la décision (V, NV, VC)"""
        try:
            decision_val = str(row.get(column_name, 'NV')).strip().upper()

            # IMPORTANT: Vérifier "NON VALIDE" AVANT "VALIDE"
            if 'NON VALIDE' in decision_val or decision_val == 'NV':
                return 'NV'
            elif 'COMPENSATION' in decision_val or decision_val == 'VC':
                return 'VC'
            elif 'VALIDE' in decision_val or decision_val == 'V':
                return 'V'
            else:
                return 'NV'
        except:
            return 'NV'
