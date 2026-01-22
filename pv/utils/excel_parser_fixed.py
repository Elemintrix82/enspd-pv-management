"""
Parser Excel FINAL - Gère correctement CC, EX, MOY, [vide], CA, [vide], DECISION
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from decimal import Decimal


class PVExcelParser:
    """Parser optimisé pour les fichiers PV ENSPD"""

    def __init__(self, file_path):
        self.file_path = file_path
        self.wb = load_workbook(file_path)
        self.ws = self.wb.active
        self.metadata = {}
        self.ues = []
        self.ecues = []
        self.etudiants = []
        self.header_row = 11

    def parse(self):
        """Parse complet"""
        self.extract_metadata()
        self.extract_structure()
        self.extract_student_data()
        return {
            'metadata': self.metadata,
            'ues': self.ues,
            'ecues': self.ecues,
            'etudiants': self.etudiants
        }

    def extract_metadata(self):
        """Extrait les métadonnées"""
        self.metadata['universite'] = self.ws['F1'].value or "UNIVERSITE DE DOUALA"
        self.metadata['ecole'] = self.ws['F3'].value or "École Nationale Supérieure Polytechnique de Douala"

        niveau_val = self.ws['I4'].value
        self.metadata['niveau'] = int(niveau_val) if niveau_val else 4

        filiere_val = self.ws['F8'].value
        if filiere_val and ':' in str(filiere_val):
            self.metadata['filiere'] = str(filiere_val).split(':', 1)[1].strip()
        else:
            self.metadata['filiere'] = str(filiere_val) if filiere_val else "GRT"

        semestre_s7 = self.ws['H7'].value
        self.metadata['semestre'] = str(semestre_s7) if semestre_s7 else "S7"

        # Année académique
        annee_found = False
        for row in range(5, 7):
            for col in range(6, 12):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and '/' in str(cell_val) and len(str(cell_val).strip()) <= 12:
                    self.metadata['annee_academique'] = str(cell_val).strip()
                    annee_found = True
                    break
            if annee_found:
                break
        if not annee_found:
            self.metadata['annee_academique'] = "2022/2023"

        # Formation
        formation_found = False
        for row in range(1, 10):
            for col in range(1, 15):
                cell_val = self.ws.cell(row=row, column=col).value
                if cell_val and 'ALTERNANCE' in str(cell_val).upper():
                    self.metadata['formation'] = "ALTERNANCE"
                    formation_found = True
                    break
                elif cell_val and 'CLASSIQUE' in str(cell_val).upper():
                    self.metadata['formation'] = "CLASSIQUE"
                    formation_found = True
                    break
            if formation_found:
                break
        if not formation_found:
            self.metadata['formation'] = "ALTERNANCE"

    def extract_structure(self):
        """Extrait UE/ECUE de la structure Excel"""
        df_structure = pd.read_excel(self.file_path, header=None, nrows=11)

        ue_row = df_structure.iloc[8]  # Ligne 9
        ecue_row = df_structure.iloc[9]  # Ligne 10

        ue_ordre = 1
        ecue_ordre = 1
        current_ue_code = None

        col_idx = 0
        while col_idx < len(ue_row):
            ue_val = ue_row.iloc[col_idx] if col_idx < len(ue_row) else None
            ecue_val = ecue_row.iloc[col_idx] if col_idx < len(ecue_row) else None

            # Nouvelle UE
            if pd.notna(ue_val) and ('EPDGIT' in str(ue_val) or 'EPDTCO' in str(ue_val)):
                parts = str(ue_val).split(':', 1)
                if len(parts) == 2:
                    ue_code = parts[0].strip()
                    ue_intitule = parts[1].strip()

                    if not any(ue['code'] == ue_code for ue in self.ues):
                        self.ues.append({
                            'code': ue_code,
                            'intitule': ue_intitule,
                            'ordre': ue_ordre
                        })
                        current_ue_code = ue_code
                        ue_ordre += 1

            # Nouvelle ECUE
            if pd.notna(ecue_val):
                ecue_str = str(ecue_val).strip()

                # ECUE normale
                if '(' in ecue_str and ')' in ecue_str and ('EPDGIT' in ecue_str or 'EPDTCO' in ecue_str):
                    code_part = ecue_str.split(')')[0]
                    ecue_code = code_part.replace('(', '').strip()
                    ecue_intitule = ecue_str.split(')', 1)[1].strip() if ')' in ecue_str else ''

                    if not any(e['code'] == ecue_code for e in self.ecues):
                        self.ecues.append({
                            'code': ecue_code,
                            'intitule': ecue_intitule,
                            'ordre': ecue_ordre,
                            'ue_code': current_ue_code,
                            'is_synthese': False
                        })
                        ecue_ordre += 1

                # SYNTHESE UE
                elif 'SYNTHESE' in ecue_str.upper() and current_ue_code:
                    synthese_code = f"SYNTHESE_{current_ue_code}"
                    if not any(e['code'] == synthese_code for e in self.ecues):
                        self.ecues.append({
                            'code': synthese_code,
                            'intitule': f"Synthèse {current_ue_code}",
                            'ordre': ecue_ordre,
                            'ue_code': current_ue_code,
                            'is_synthese': True
                        })
                        ecue_ordre += 1

            col_idx += 1

    def extract_student_data(self):
        """Extrait les données des étudiants"""
        df = pd.read_excel(self.file_path, header=self.header_row - 1)
        df.columns = [str(col).strip() if pd.notna(col) else f"Unnamed_{i}"
                      for i, col in enumerate(df.columns)]

        for idx, row in df.iterrows():
            numero = row.get('N°', idx + 1)
            matricule = str(row.get('MATRICULE', '')).strip()
            nom_prenom = str(row.get('NOMS & PRENOMS', '')).strip()

            if not matricule or matricule == 'nan':
                continue

            moyenne_generale = self._safe_decimal(row.get('MOYENNE/20', 0))
            credits_acquis = self._safe_int(row.get('CREDITS  ACQUIS', 0))
            decision_generale = self._extract_decision(row, df.columns[-1])

            # Extraire notes et synthèses
            notes_par_ecue = []
            syntheses_par_ue = []

            # Stratégie: parcourir TOUTES les colonnes CC et associer avec les ECUE par ordre
            ecue_notes = []
            synthese_ues = []

            i = 0
            while i < len(df.columns):
                col_str = str(df.columns[i]).upper()

                # Détecter séquence ECUE: CC, EX, MOY, [vide], CA, [vide], DECISION
                if ('CC' == col_str or 'CC.' in col_str) and i + 6 < len(df.columns):
                    next_cols = [str(df.columns[i+j]).upper() for j in range(1, 7)]

                    # Vérifier: EX, MOY, ?, CA, ?, DECISION
                    if 'EX' in next_cols[0] and 'MOY' in next_cols[1] and 'CA' in next_cols[3] and 'DECISION' in next_cols[5]:
                        cc = self._safe_decimal(row.iloc[i])
                        ex = self._safe_decimal(row.iloc[i+1])
                        moy = self._safe_decimal(row.iloc[i+2])
                        ca = self._safe_int(row.iloc[i+4])
                        dec = self._extract_decision(row, df.columns[i+6])

                        if moy > 0:
                            ecue_notes.append({
                                'cc': cc,
                                'examen': ex,
                                'moyenne': moy,
                                'credit_attribue': ca,
                                'decision': dec
                            })
                        i += 7  # Sauter toute la séquence
                        continue

                # Détecter séquence SYNTHESE UE: MOY, [vide], CA, [vide], DECISION (sans CC/EX avant)
                if i > 0 and ('MOY' in col_str) and i + 4 < len(df.columns):
                    prev_col = str(df.columns[i-1]).upper()
                    # Vérifier qu'il n'y a pas CC/EX juste avant
                    if 'CC' not in prev_col and 'EX' not in prev_col:
                        next_cols = [str(df.columns[i+j]).upper() for j in range(1, 5)]

                        # Vérifier: ?, CA, ?, DECISION
                        if ('CA' in next_cols[1] or 'UNNAMED' in next_cols[1]) and 'DECISION' in next_cols[3]:
                            moy = self._safe_decimal(row.iloc[i])
                            ca = self._safe_int(row.iloc[i+2])
                            dec = self._extract_decision(row, df.columns[i+4])

                            if moy > 0:
                                synthese_ues.append({
                                    'moyenne_ue': moy,
                                    'credits_attribues': ca,
                                    'decision': dec
                                })
                            i += 5
                            continue

                i += 1

            # Associer les notes extraites aux ECUE par ordre
            ecue_reelles = [e for e in self.ecues if not e.get('is_synthese', False)]
            for idx_note, note in enumerate(ecue_notes):
                if idx_note < len(ecue_reelles):
                    note['ecue_code'] = ecue_reelles[idx_note]['code']
                    notes_par_ecue.append(note)

            # Associer les synthèses aux UE par ordre
            for idx_synth, synthese in enumerate(synthese_ues):
                if idx_synth < len(self.ues):
                    synthese['ue_code'] = self.ues[idx_synth]['code']
                    syntheses_par_ue.append(synthese)

            etudiant_data = {
                'numero': self._safe_int(numero),
                'matricule': matricule,
                'nom_prenom': nom_prenom,
                'moyenne_generale': moyenne_generale,
                'credits_acquis': credits_acquis,
                'decision_generale': decision_generale,
                'notes': notes_par_ecue,
                'syntheses_ue': syntheses_par_ue
            }

            self.etudiants.append(etudiant_data)

    def _safe_decimal(self, value):
        """Convertit en Decimal"""
        try:
            if pd.isna(value):
                return Decimal('0.00')
            return Decimal(str(value))
        except:
            return Decimal('0.00')

    def _safe_int(self, value):
        """Convertit en int"""
        try:
            if pd.isna(value):
                return 0
            return int(float(value))
        except:
            return 0

    def _extract_decision(self, row, column_name):
        """Extrait la décision"""
        try:
            decision_val = str(row.get(column_name, 'NV')).strip().upper()
            if 'NON VALIDE' in decision_val or decision_val == 'NV':
                return 'NV'
            elif 'COMPENSATION' in decision_val or decision_val == 'VC':
                return 'VC'
            elif 'VALIDE' in decision_val or decision_val == 'V':
                return 'V'
            else:
                return 'NV'
        except:
            return 'NV'
