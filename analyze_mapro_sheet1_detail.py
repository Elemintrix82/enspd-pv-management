"""
Analyse détaillée de la première feuille de MAPRO pour comprendre sa structure
"""
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

mapro_path = r"D:\RYDI_Group\ENSPD\MAPRO_GIT5_SN_SEM1.xlsx"
wb = load_workbook(mapro_path, data_only=True)

# Analyser la première feuille
ws = wb['PV_M2PDGL2_02_FEV_14']

print("=" * 80)
print("ANALYSE DÉTAILLÉE - FEUILLE: PV_M2PDGL2_02_FEV_14")
print("=" * 80)
print()

# Afficher les 15 premières lignes complètes
print("PREMIÈRES 15 LIGNES (colonnes A-O):")
print("=" * 80)

for row in range(1, 16):
    print(f"\n[Ligne {row}]")
    for col in range(1, 16):  # Colonnes A-O
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            col_letter = get_column_letter(col)
            value_str = str(cell_value)[:60]
            print(f"  {col_letter}: {value_str}")

# Chercher spécifiquement les codes MPGIT
print("\n" + "=" * 80)
print("RECHERCHE DES CODES MPGIT DANS LES LIGNES 8-11:")
print("=" * 80)

for row in range(8, 12):
    print(f"\n[Ligne {row}]")
    has_code = False
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            cell_str = str(cell_value)
            if 'MPGIT' in cell_str or 'MPG' in cell_str or '(' in cell_str:
                col_letter = get_column_letter(col)
                print(f"  {col_letter}: {cell_str}")
                has_code = True
    if not has_code:
        print("  (Aucun code détecté sur cette ligne)")

print("\n" + "=" * 80)
print("SCAN COMPLET DES 20 PREMIÈRES COLONNES (lignes 1-15):")
print("=" * 80)

for col in range(1, 21):
    col_letter = get_column_letter(col)
    print(f"\n[Colonne {col_letter}]")
    for row in range(1, 16):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value and len(str(cell_value).strip()) > 1:
            value_str = str(cell_value)[:50]
            print(f"  Ligne {row}: {value_str}")
