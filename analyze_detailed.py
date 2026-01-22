"""
Analyse detaillee de la structure hierarchique du fichier PV Excel
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook
import json

file_path = r"D:\RYDI_Group\ENSPD\PV_GRT4_SEM7_ALT.xlsx"

print("=" * 80)
print("ANALYSE DETAILLEE DE LA STRUCTURE HIERARCHIQUE")
print("=" * 80)

# Charger le workbook avec openpyxl pour acceder aux cellules fusionnees
wb = load_workbook(file_path)
ws = wb.active

# 1. Lire les lignes de metadonnees (lignes 1-10)
print("\n1. METADONNEES EXTRAITES")
print("-" * 80)

metadata = {
    'universite': ws['F1'].value,
    'ecole': ws['F3'].value,
    'niveau': ws['I4'].value,
    'filiere': ws['F8'].value,
    'semestre_s7': ws['H7'].value,
    'semestre_s8': ws['I7'].value,
    'annee_academique_1': ws['H6'].value,
    'annee_academique_2': ws['I6'].value,
    'formation': ws['I6'].value,
}

for key, value in metadata.items():
    print(f"{key}: {value}")

# 2. Identifier les UE et ECUE en analysant les lignes 8-9
print("\n2. STRUCTURE HIERARCHIQUE (UE -> ECUE)")
print("-" * 80)

# Lire ligne 9 (index 8) pour les UE
ue_row = 8
ecue_row = 9
header_row = 10

# Structure pour stocker la hierarchie
structure = {
    'ue_list': [],
    'ecue_list': [],
    'mapping': {}
}

# Analyser les UE (ligne 9)
print("\nUNITES D'ENSEIGNEMENT (UE):")
current_col = 5  # Colonne F (index 5)
ue_index = 1

while current_col < ws.max_column:
    cell_value = ws.cell(row=ue_row+1, column=current_col+1).value
    if cell_value and 'EPDGIT' in str(cell_value) or 'EPDTCO' in str(cell_value):
        # Extraire code et intitule
        parts = str(cell_value).split(':', 1)
        if len(parts) == 2:
            ue_code = parts[0].strip()
            ue_intitule = parts[1].strip()
            structure['ue_list'].append({
                'ordre': ue_index,
                'code': ue_code,
                'intitule': ue_intitule,
                'colonne_debut': current_col
            })
            print(f"{ue_index}. {ue_code}: {ue_intitule}")
            ue_index += 1
    current_col += 1

# Analyser les ECUE (ligne 10)
print("\nELEMENTS CONSTITUTIFS (ECUE):")
current_col = 5
ecue_index = 1

while current_col < ws.max_column:
    cell_value = ws.cell(row=ecue_row+1, column=current_col+1).value
    if cell_value and 'EPDGIT' in str(cell_value) or 'EPDTCO' in str(cell_value):
        # Extraire code et intitule
        if ')' in str(cell_value):
            code_part = str(cell_value).split(')')[0]
            code = code_part.replace('(', '').strip()
            intitule = str(cell_value).split(')', 1)[1].strip() if ')' in str(cell_value) else ''

            structure['ecue_list'].append({
                'ordre': ecue_index,
                'code': code,
                'intitule': intitule,
                'colonne_debut': current_col
            })
            print(f"{ecue_index}. {code}: {intitule}")
            ecue_index += 1
    current_col += 1

# 3. Analyser la ligne d'en-tete (ligne 11, index 10)
print("\n3. STRUCTURE DES COLONNES DE NOTES")
print("-" * 80)

# Lire avec pandas
df = pd.read_excel(file_path, header=10)
print(f"Nombre total de colonnes: {len(df.columns)}")

# Identifier les colonnes identite
colonnes_identite = ['N°', 'MATRICULE', 'NOMS & PRENOMS']
print(f"\nColonnes d'identite: {colonnes_identite}")

# Analyser le pattern repetitif: CC, EX, MOY, CA, DECISION
print("\nPattern de colonnes pour chaque ECUE:")
print("  - CC (Controle Continu)")
print("  - EX (Examen)")
print("  - MOY (Moyenne)")
print("  - CA (Credits Attribues)")
print("  - DECISION (V/NV/VC)")

# 4. Compter les etudiants
print("\n4. STATISTIQUES ETUDIANTS")
print("-" * 80)
print(f"Nombre total d'etudiants: {len(df)}")

# Analyser les decisions finales
decision_finale_col = 'DECISION.16'  # Derniere colonne DECISION
if decision_finale_col in df.columns:
    print("\nRepartition des decisions finales:")
    decisions = df[decision_finale_col].value_counts()
    for decision, count in decisions.items():
        pourcentage = (count / len(df)) * 100
        print(f"  {decision}: {count} ({pourcentage:.1f}%)")

# 5. Analyser les moyennes
print("\n5. STATISTIQUES DES MOYENNES")
print("-" * 80)
moyenne_col = 'MOYENNE/20'
if moyenne_col in df.columns:
    print(f"Moyenne generale:")
    print(f"  Min: {df[moyenne_col].min():.2f}")
    print(f"  Max: {df[moyenne_col].max():.2f}")
    print(f"  Moyenne: {df[moyenne_col].mean():.2f}")
    print(f"  Mediane: {df[moyenne_col].median():.2f}")

# 6. Creer un mapping des colonnes
print("\n6. MAPPING COMPLET DES COLONNES")
print("-" * 80)

column_mapping = {}
col_index = 0

for col_name in df.columns:
    column_mapping[col_index] = {
        'nom': col_name,
        'index': col_index,
        'type': 'identite' if col_name in colonnes_identite else 'note'
    }
    col_index += 1

# Sauvegarder le mapping
with open(r'D:\RYDI_Group\ENSPD\column_mapping.json', 'w', encoding='utf-8') as f:
    json.dump(column_mapping, f, ensure_ascii=False, indent=2)

print("Mapping sauvegarde dans: column_mapping.json")

# 7. Creer un document de mapping final
print("\n7. CREATION DU DOCUMENT DE MAPPING")
print("-" * 80)

mapping_doc = f"""
DOCUMENT DE MAPPING - FICHIER PV EXCEL
================================================================================

METADONNEES
-----------
Universite: {metadata['universite']}
Ecole: {metadata['ecole']}
Filiere: {metadata['filiere']}
Niveau: {metadata['niveau']}
Semestre: {metadata['semestre_s7']}

STRUCTURE DU FICHIER
--------------------
Ligne 1-7: Metadonnees institutionnelles
Ligne 8-9: Titres UE et ECUE
Ligne 10: En-tetes de colonnes (N°, MATRICULE, NOMS & PRENOMS, puis notes)
Ligne 11-38: Donnees etudiants (28 etudiants)

UNITES D'ENSEIGNEMENT (UE)
--------------------------
{len(structure['ue_list'])} UE identifiees:
"""

for ue in structure['ue_list']:
    mapping_doc += f"{ue['ordre']}. {ue['code']} - {ue['intitule']}\n"

mapping_doc += f"""

ELEMENTS CONSTITUTIFS (ECUE)
----------------------------
{len(structure['ecue_list'])} ECUE identifiees:
"""

for ecue in structure['ecue_list']:
    mapping_doc += f"{ecue['ordre']}. {ecue['code']} - {ecue['intitule']}\n"

mapping_doc += """

STRUCTURE DES COLONNES
----------------------
Colonnes 1-3: Identite (N°, MATRICULE, NOMS & PRENOMS)
Colonnes suivantes: Pattern repetitif pour chaque ECUE
  - CC (Controle Continu)
  - EX (Examen)
  - MOY (Moyenne)
  - CA (Credits Attribues)
  - DECISION (V=Valide, NV=Non Valide, VC=Valide par Compensation)

Apres toutes les ECUE, syntheses par UE, puis:
  - MOYENNE/20 (Moyenne generale)
  - CREDITS ACQUIS (Total credits)
  - DECISION (Decision finale)

VALEURS DE DECISION
-------------------
V  = Valide
NV = Non Valide
VC = Valide par Compensation

STATISTIQUES
------------
Total etudiants: {len(df)}
Nombre de colonnes: {len(df.columns)}
"""

# Sauvegarder le document
with open(r'D:\RYDI_Group\ENSPD\MAPPING_COMPLET.txt', 'w', encoding='utf-8') as f:
    f.write(mapping_doc)

print("Document de mapping sauvegarde dans: MAPPING_COMPLET.txt")

# 8. Afficher un resume
print("\n" + "=" * 80)
print("RESUME DE L'ANALYSE")
print("=" * 80)
print(f"Total UE: {len(structure['ue_list'])}")
print(f"Total ECUE: {len(structure['ecue_list'])}")
print(f"Total Etudiants: {len(df)}")
print(f"Total Colonnes: {len(df.columns)}")
print("\nFichiers generes:")
print("  - column_mapping.json")
print("  - MAPPING_COMPLET.txt")
print("\n" + "=" * 80)
