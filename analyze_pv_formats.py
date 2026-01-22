"""
Script d'analyse automatique des formats de PV
Analyse la structure de chaque fichier Excel pour détecter les différences
et préparer l'adaptation du système

Date: 17 janvier 2026
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from pathlib import Path
from datetime import datetime


class PVFormatAnalyzer:
    """Analyseur de structure de fichiers PV Excel"""

    def __init__(self, excel_path):
        self.excel_path = Path(excel_path)
        self.wb = load_workbook(excel_path, data_only=True)
        self.ws = self.wb.active
        self.analysis = {
            'filename': self.excel_path.name,
            'filepath': str(self.excel_path),
            'analyzed_at': datetime.now().isoformat(),
        }

    def analyze(self):
        """Analyse complète du fichier"""
        print(f"\n{'='*80}")
        print(f"ANALYSE : {self.excel_path.name}")
        print(f"{'='*80}\n")

        self.analyze_general_info()
        self.analyze_header_structure()
        self.analyze_fixed_columns()
        self.analyze_ue_structure()
        self.analyze_data_area()
        self.analyze_nomenclature()
        self.analyze_particularities()

        return self.analysis

    def analyze_general_info(self):
        """Informations générales du fichier"""
        print("[INFORMATIONS GENERALES]")
        print("-" * 80)

        info = {
            'sheet_count': len(self.wb.sheetnames),
            'sheet_names': self.wb.sheetnames,
            'active_sheet': self.wb.active.title,
            'max_row': self.ws.max_row,
            'max_column': self.ws.max_column,
            'max_column_letter': get_column_letter(self.ws.max_column),
        }

        for key, value in info.items():
            print(f"  {key}: {value}")

        self.analysis['general_info'] = info
        print()

    def analyze_header_structure(self):
        """Analyse de la structure des en-têtes"""
        print("[STRUCTURE DES EN-TETES]")
        print("-" * 80)

        # Détecter le nombre de lignes d'en-têtes
        header_rows = self.detect_header_rows()
        data_start_row = header_rows + 1

        header_structure = {
            'header_rows_count': header_rows,
            'data_start_row': data_start_row,
            'headers': {}
        }

        # Extraire les en-têtes ligne par ligne
        for row_num in range(1, header_rows + 1):
            row_headers = []
            for col in range(1, self.ws.max_column + 1):
                cell = self.ws.cell(row_num, col)
                value = cell.value if cell.value is not None else ""
                row_headers.append({
                    'column': get_column_letter(col),
                    'value': str(value),
                    'merged': isinstance(cell, openpyxl.cell.cell.MergedCell)
                })
            header_structure['headers'][f'row_{row_num}'] = row_headers

        print(f"  Nombre de lignes d'en-têtes: {header_rows}")
        print(f"  Ligne de début des données: {data_start_row}")
        print()

        self.analysis['header_structure'] = header_structure

    def detect_header_rows(self):
        """Détecte le nombre de lignes d'en-têtes"""
        # Chercher la première ligne contenant un matricule ou un numéro d'étudiant
        for row in range(1, 15):
            cell_value = self.ws.cell(row, 2).value  # Colonne B (probablement matricule)
            if cell_value and isinstance(cell_value, str):
                # Si on trouve un matricule typique (ex: "24G01854", "22A001")
                if any(char.isdigit() for char in str(cell_value)) and len(str(cell_value)) >= 5:
                    return row - 1

        # Valeur par défaut si détection échoue
        return 3

    def analyze_fixed_columns(self):
        """Analyse des colonnes fixes (N°, Matricule, Nom)"""
        print("[*] COLONNES FIXES")
        print("-" * 80)

        # Rechercher les colonnes fixes dans la première ligne d'en-tête
        fixed_cols = []
        keywords = ['N°', 'MATRICULE', 'NOM', 'PRÉNOM', 'PRENOM']

        for col in range(1, min(10, self.ws.max_column + 1)):  # Chercher dans les 10 premières colonnes
            for row in range(1, 4):  # Chercher dans les 3 premières lignes
                cell_value = str(self.ws.cell(row, col).value or "").strip().upper()
                for keyword in keywords:
                    if keyword in cell_value:
                        fixed_cols.append({
                            'column': get_column_letter(col),
                            'column_index': col,
                            'label': cell_value,
                            'keyword': keyword,
                            'row': row
                        })
                        break

        # Supprimer les doublons
        seen = set()
        unique_fixed_cols = []
        for col in fixed_cols:
            if col['column'] not in seen:
                seen.add(col['column'])
                unique_fixed_cols.append(col)
                print(f"  {col['column']}: {col['label']}")

        self.analysis['fixed_columns'] = unique_fixed_cols
        print()

    def analyze_ue_structure(self):
        """Analyse de la structure des UE et ECUE"""
        print("[*] STRUCTURE UE/ECUE")
        print("-" * 80)

        # Chercher les codes UE dans la première ligne d'en-tête
        ue_list = []
        ecue_list = []

        # Ligne 1: Codes UE (généralement à partir de la colonne D ou E)
        start_col = 4  # Après les colonnes fixes

        for col in range(start_col, self.ws.max_column + 1):
            cell_value = str(self.ws.cell(1, col).value or "").strip()

            # Détecter les codes UE (ex: EPD...)
            if cell_value and cell_value.startswith('EPD'):
                ue_list.append({
                    'column': get_column_letter(col),
                    'column_index': col,
                    'code': cell_value,
                })

        # Ligne 2: Codes ECUE
        for col in range(start_col, self.ws.max_column + 1):
            cell_value = str(self.ws.cell(2, col).value or "").strip()

            # Détecter les codes ECUE
            if cell_value and (cell_value.startswith('EPD') or cell_value.startswith('MAPRO')):
                ecue_list.append({
                    'column': get_column_letter(col),
                    'column_index': col,
                    'code': cell_value,
                })

        print(f"  Nombre d'UE détectées: {len(ue_list)}")
        if ue_list:
            print(f"  Codes UE: {', '.join([ue['code'] for ue in ue_list[:5]])}{'...' if len(ue_list) > 5 else ''}")

        print(f"  Nombre d'ECUE détectés: {len(ecue_list)}")
        if ecue_list:
            print(f"  Codes ECUE: {', '.join([ecue['code'] for ecue in ecue_list[:5]])}{'...' if len(ecue_list) > 5 else ''}")

        self.analysis['ue_structure'] = {
            'ue_count': len(ue_list),
            'ue_list': ue_list,
            'ecue_count': len(ecue_list),
            'ecue_list': ecue_list,
        }
        print()

    def analyze_data_area(self):
        """Analyse de la zone de données"""
        print("[*] ZONE DE DONNÉES")
        print("-" * 80)

        data_start_row = self.analysis['header_structure']['data_start_row']

        # Compter les étudiants
        student_count = 0
        for row in range(data_start_row, self.ws.max_row + 1):
            # Vérifier si la ligne contient un matricule (colonne B)
            matricule = self.ws.cell(row, 2).value
            if matricule and str(matricule).strip():
                student_count += 1

        data_info = {
            'student_count': student_count,
            'data_start_row': data_start_row,
            'data_end_row': data_start_row + student_count - 1,
        }

        print(f"  Ligne de début des données: {data_start_row}")
        print(f"  Nombre d'étudiants: {student_count}")
        print(f"  Ligne de fin des données: {data_info['data_end_row']}")

        self.analysis['data_area'] = data_info
        print()

    def analyze_nomenclature(self):
        """Analyse de la nomenclature (codes, matricules, décisions)"""
        print("[*]  NOMENCLATURE")
        print("-" * 80)

        data_start_row = self.analysis['header_structure']['data_start_row']

        # Échantillonner quelques matricules
        matricules_sample = []
        for row in range(data_start_row, min(data_start_row + 5, self.ws.max_row + 1)):
            matricule = self.ws.cell(row, 2).value
            if matricule:
                matricules_sample.append(str(matricule))

        # Chercher des décisions (V, NV, VC)
        decisions_found = set()
        for row in range(data_start_row, min(data_start_row + 10, self.ws.max_row + 1)):
            for col in range(1, self.ws.max_column + 1):
                value = self.ws.cell(row, col).value
                if value in ['V', 'NV', 'VC', 'ABI', 'EXC']:
                    decisions_found.add(value)

        nomenclature = {
            'matricules_sample': matricules_sample,
            'matricule_format': self.detect_matricule_format(matricules_sample),
            'decisions_found': list(decisions_found),
        }

        print(f"  Format matricules: {nomenclature['matricule_format']}")
        print(f"  Exemples de matricules: {', '.join(matricules_sample[:3])}")
        print(f"  Décisions trouvées: {', '.join(decisions_found) if decisions_found else 'Aucune'}")

        self.analysis['nomenclature'] = nomenclature
        print()

    def detect_matricule_format(self, matricules):
        """Détecte le format des matricules"""
        if not matricules:
            return "UNKNOWN"

        sample = matricules[0]

        # Format type: 24G01854 (2 chiffres + 1 lettre + 5 chiffres)
        if len(sample) == 8 and sample[0:2].isdigit() and sample[2].isalpha() and sample[3:].isdigit():
            return "YYXNNNNN (ex: 24G01854)"

        # Format type: 22A001 (2 chiffres + 1 lettre + 3 chiffres)
        if len(sample) == 6 and sample[0:2].isdigit() and sample[2].isalpha() and sample[3:].isdigit():
            return "YYXNNN (ex: 22A001)"

        return f"CUSTOM (longueur: {len(sample)})"

    def analyze_particularities(self):
        """Analyse des particularités du fichier"""
        print("[!]  PARTICULARITÉS")
        print("-" * 80)

        particularities = {
            'has_merged_cells': False,
            'has_formulas': False,
            'has_empty_cells': False,
            'has_special_values': False,
            'special_values_found': [],
        }

        # Vérifier les cellules fusionnées
        if len(self.ws.merged_cells.ranges) > 0:
            particularities['has_merged_cells'] = True
            print(f"  [!]  Cellules fusionnées détectées: {len(self.ws.merged_cells.ranges)}")

        # Vérifier les formules
        for row in self.ws.iter_rows(min_row=1, max_row=min(10, self.ws.max_row)):
            for cell in row:
                if cell.data_type == 'f':
                    particularities['has_formulas'] = True
                    break

        if particularities['has_formulas']:
            print(f"  [!]  Formules Excel détectées")

        # Vérifier les valeurs spéciales
        data_start_row = self.analysis['header_structure']['data_start_row']
        special_values = ['---', 'ABI', 'EXC', 'N/A', 'NULL']

        for row in range(data_start_row, min(data_start_row + 20, self.ws.max_row + 1)):
            for col in range(1, self.ws.max_column + 1):
                value = self.ws.cell(row, col).value
                if value in special_values:
                    particularities['has_special_values'] = True
                    if value not in particularities['special_values_found']:
                        particularities['special_values_found'].append(value)

        if particularities['special_values_found']:
            print(f"  [!]  Valeurs spéciales: {', '.join(particularities['special_values_found'])}")

        if not any(particularities.values()):
            print(f"  [OK] Aucune particularité détectée")

        self.analysis['particularities'] = particularities
        print()

    def generate_summary(self):
        """Génère un résumé de l'analyse"""
        print("[*] RÉSUMÉ DE L'ANALYSE")
        print("=" * 80)

        summary = {
            'header_rows': self.analysis['header_structure']['header_rows_count'],
            'data_start_row': self.analysis['header_structure']['data_start_row'],
            'fixed_columns_count': len(self.analysis['fixed_columns']),
            'ue_count': self.analysis['ue_structure']['ue_count'],
            'ecue_count': self.analysis['ue_structure']['ecue_count'],
            'student_count': self.analysis['data_area']['student_count'],
            'total_columns': self.analysis['general_info']['max_column'],
        }

        for key, value in summary.items():
            print(f"  {key}: {value}")

        self.analysis['summary'] = summary
        print()

        return summary

    def save_report(self, output_dir="analysis_reports"):
        """Sauvegarde le rapport d'analyse en JSON"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)

        report_file = output_path / f"{self.excel_path.stem}_analysis.json"

        with open(report_file, 'w', encoding='utf-8') as f:
            json.dump(self.analysis, f, indent=2, ensure_ascii=False)

        print(f"[OK] Rapport sauvegardé: {report_file}")

        return report_file


def analyze_all_pv_files():
    """Analyse tous les fichiers PV disponibles"""

    # Liste des fichiers à analyser
    files_to_analyze = [
        "D:\\RYDI_Group\\ENSPD\\media\\pv\\PV_GRT4_SEM7_ALT.xlsx",  # Référence
        "D:\\RYDI_Group\\ENSPD\\MAPRO_GIT5_SN_SEM1.xlsx",
        "D:\\RYDI_Group\\ENSPD\\PV_GL04_SEM7_ALT.xlsx",
        "D:\\RYDI_Group\\ENSPD\\PV_GL04_SEM7_FI1.xlsx",
        "D:\\RYDI_Group\\ENSPD\\PV_GLO5.xlsx",
        "D:\\RYDI_Group\\ENSPD\\media\\pv\\PV_GRT4_SEM7_FI1.xlsx",
        "D:\\RYDI_Group\\ENSPD\\media\\pv\\PV_GRT5_SEM9_FI1.xlsx",
    ]

    results = []

    for file_path in files_to_analyze:
        if not Path(file_path).exists():
            print(f"[!]  Fichier introuvable: {file_path}")
            continue

        try:
            analyzer = PVFormatAnalyzer(file_path)
            analysis = analyzer.analyze()
            summary = analyzer.generate_summary()
            analyzer.save_report()

            results.append({
                'filename': Path(file_path).name,
                'status': 'SUCCESS',
                'summary': summary
            })

        except Exception as e:
            print(f"[X] Erreur lors de l'analyse de {file_path}: {e}")
            results.append({
                'filename': Path(file_path).name,
                'status': 'ERROR',
                'error': str(e)
            })

    return results


def generate_compatibility_table(results):
    """Génère le tableau de compatibilité"""
    print("\n" + "=" * 80)
    print("TABLEAU DE COMPATIBILITÉ")
    print("=" * 80 + "\n")

    # Header
    print(f"{'Fichier':<40} {'En-têtes':<12} {'UE':<8} {'ECUE':<8} {'Étudiants':<12}")
    print("-" * 80)

    # Référence
    reference = None
    for result in results:
        if 'PV_GRT4_SEM7_ALT' in result['filename']:
            reference = result
            break

    if reference:
        ref_summary = reference['summary']
        print(f"{'[REF] ' + reference['filename']:<40} "
              f"{ref_summary['header_rows']:<12} "
              f"{ref_summary['ue_count']:<8} "
              f"{ref_summary['ecue_count']:<8} "
              f"{ref_summary['student_count']:<12}")
        print("-" * 80)

    # Autres fichiers
    for result in results:
        if result['filename'] == reference['filename']:
            continue

        if result['status'] == 'SUCCESS':
            summary = result['summary']

            # Comparer avec la référence
            header_match = "[OK]" if summary['header_rows'] == ref_summary['header_rows'] else "[!]"

            print(f"{result['filename']:<40} "
                  f"{header_match} {summary['header_rows']:<10} "
                  f"{summary['ue_count']:<8} "
                  f"{summary['ecue_count']:<8} "
                  f"{summary['student_count']:<12}")
        else:
            print(f"{result['filename']:<40} [X] ERREUR")

    print()


if __name__ == "__main__":
    print("\n" + "=" * 80)
    print("ANALYSEUR AUTOMATIQUE DE FORMATS PV - ENSPD")
    print("Analyse structurelle de tous les fichiers Excel")
    print("Date: 17 janvier 2026")
    print("=" * 80 + "\n")

    results = analyze_all_pv_files()
    generate_compatibility_table(results)

    print("\n[OK] Analyse terminee !")
    print(f"[*] Rapports sauvegardes dans: analysis_reports/")
