"""
Script d'analyse du fichier PV Excel
Analyse complète de la structure du fichier PV_GRT4_SEM7_ALT.xlsx
"""

import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Chemin du fichier
file_path = r"D:\RYDI_Group\ENSPD\PV_GRT4_SEM7_ALT.xlsx"

print("=" * 80)
print("ANALYSE COMPLÈTE DU FICHIER PV EXCEL")
print("=" * 80)

# 1. Informations générales du fichier
print("\n1. INFORMATIONS GÉNÉRALES")
print("-" * 80)
wb = load_workbook(file_path)
print(f"Nombre de feuilles: {len(wb.sheetnames)}")
print(f"Noms des feuilles: {wb.sheetnames}")

# Travailler avec la première feuille
ws = wb.active
print(f"Feuille active: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Nombre de lignes: {ws.max_row}")
print(f"Nombre de colonnes: {ws.max_column}")

# 2. Lecture des premières lignes pour identifier les métadonnées
print("\n2. MÉTADONNÉES ET EN-TÊTES (20 premières lignes)")
print("-" * 80)
df_raw = pd.read_excel(file_path, header=None, nrows=20)
for idx, row in df_raw.iterrows():
    print(f"Ligne {idx}: {row.tolist()[:10]}...")  # Afficher les 10 premières valeurs

# 3. Identifier la ligne d'en-tête des colonnes
print("\n3. IDENTIFICATION DE LA LIGNE D'EN-TÊTE")
print("-" * 80)
for idx in range(20):
    row_values = df_raw.iloc[idx].tolist()
    if any(str(val).strip().upper() in ['N°', 'MATRICULE', 'NOM', 'PRENOM'] for val in row_values if pd.notna(val)):
        print(f"Ligne d'en-tête probable: Ligne {idx}")
        print(f"Valeurs: {row_values}")
        header_row = idx
        break

# 4. Lire le fichier avec l'en-tête identifié
print("\n4. STRUCTURE DES COLONNES")
print("-" * 80)
try:
    # Essayer de lire avec différentes lignes d'en-tête
    for skip in range(15):
        try:
            df = pd.read_excel(file_path, header=skip)
            cols = df.columns.tolist()
            if len(cols) > 5 and any('matricule' in str(col).lower() for col in cols if pd.notna(col)):
                print(f"\n✓ En-tête trouvé à la ligne {skip}")
                print(f"Nombre de colonnes: {len(cols)}")
                print(f"\nListe complète des colonnes:")
                for i, col in enumerate(cols, 1):
                    print(f"  {i}. {col}")
                header_row = skip
                break
        except:
            continue
except Exception as e:
    print(f"Erreur: {e}")

# 5. Analyser les données étudiants
print("\n5. ANALYSE DES DONNÉES ÉTUDIANTS")
print("-" * 80)
df = pd.read_excel(file_path, header=header_row)
print(f"Nombre total d'étudiants: {len(df)}")
print(f"\nAperçu des 5 premières lignes:")
print(df.head())

# 6. Identifier les colonnes par type
print("\n6. CLASSIFICATION DES COLONNES")
print("-" * 80)

colonnes_identite = []
colonnes_ue = []
colonnes_ecue = []
colonnes_notes = []

for col in df.columns:
    col_str = str(col).upper()
    if any(keyword in col_str for keyword in ['N°', 'MATRICULE', 'NOM', 'PRENOM']):
        colonnes_identite.append(col)
    elif 'UE' in col_str or 'UNITE' in col_str:
        colonnes_ue.append(col)
    elif 'ECUE' in col_str:
        colonnes_ecue.append(col)
    elif any(keyword in col_str for keyword in ['CC', 'EX', 'MOY', 'CA', 'DECISION', 'CREDIT']):
        colonnes_notes.append(col)

print(f"\nColonnes d'identité ({len(colonnes_identite)}):")
for col in colonnes_identite:
    print(f"  - {col}")

print(f"\nColonnes UE ({len(colonnes_ue)}):")
for col in colonnes_ue:
    print(f"  - {col}")

print(f"\nColonnes ECUE ({len(colonnes_ecue)}):")
for col in colonnes_ecue:
    print(f"  - {col}")

print(f"\nColonnes Notes/Décisions ({len(colonnes_notes)}):")
for col in colonnes_notes:
    print(f"  - {col}")

# 7. Analyser les valeurs de décision
print("\n7. ANALYSE DES DÉCISIONS")
print("-" * 80)
decision_cols = [col for col in df.columns if 'DECISION' in str(col).upper()]
print(f"Colonnes de décision trouvées: {len(decision_cols)}")

for col in decision_cols[:5]:  # Afficher les 5 premières
    unique_vals = df[col].dropna().unique()
    print(f"\n{col}:")
    print(f"  Valeurs uniques: {unique_vals}")
    print(f"  Répartition:")
    print(df[col].value_counts())

# 8. Extraction des métadonnées
print("\n8. EXTRACTION DES MÉTADONNÉES")
print("-" * 80)
metadata = {}
for idx in range(min(header_row, 15)):
    row_values = df_raw.iloc[idx].tolist()
    row_str = ' '.join([str(v) for v in row_values if pd.notna(v)])

    if 'filiere' in row_str.lower() or 'filière' in row_str.lower():
        metadata['filiere'] = row_str
    elif 'niveau' in row_str.lower():
        metadata['niveau'] = row_str
    elif 'semestre' in row_str.lower():
        metadata['semestre'] = row_str
    elif 'annee' in row_str.lower() or 'année' in row_str.lower():
        metadata['annee_academique'] = row_str
    elif 'formation' in row_str.lower():
        metadata['formation'] = row_str

print("Métadonnées identifiées:")
for key, value in metadata.items():
    print(f"  {key}: {value}")

# 9. Structure hiérarchique UE -> ECUE
print("\n9. STRUCTURE HIÉRARCHIQUE (UE → ECUE)")
print("-" * 80)

# Analyse des patterns de colonnes
print("Analyse du pattern des colonnes pour identifier la hiérarchie...")

# Chercher les patterns répétitifs (CC, EX, MOY, CA, DECISION)
pattern_sequence = []
current_ue = None
current_ecue = None

for col in df.columns:
    col_str = str(col)
    print(f"\nAnalyse: {col_str}")

# 10. Statistiques générales
print("\n10. STATISTIQUES GÉNÉRALES")
print("-" * 80)
print(f"Nombre total d'étudiants: {len(df)}")
print(f"Nombre de colonnes: {len(df.columns)}")
print(f"Taux de remplissage moyen: {(1 - df.isnull().sum().sum() / (df.shape[0] * df.shape[1])) * 100:.2f}%")

# 11. Sauvegarde de l'analyse
print("\n11. SAUVEGARDE DE L'ANALYSE")
print("-" * 80)

# Sauvegarder la structure des colonnes
with open('D:\\RYDI_Group\\ENSPD\\analyse_structure.txt', 'w', encoding='utf-8') as f:
    f.write("STRUCTURE DES COLONNES DU FICHIER PV\n")
    f.write("=" * 80 + "\n\n")
    f.write(f"Nombre total de colonnes: {len(df.columns)}\n\n")
    for i, col in enumerate(df.columns, 1):
        f.write(f"{i}. {col}\n")

print("✓ Analyse sauvegardée dans: analyse_structure.txt")

print("\n" + "=" * 80)
print("ANALYSE TERMINÉE")
print("=" * 80)
